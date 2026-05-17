"""
Cadence — a content pipeline for LinkedIn

A self-hosted Flask app that turns posting into a repeatable pipeline:
    Ideas  →  Drafts  →  Scheduled  →  Published  →  Analytics  →  Repurpose

All AI generation is powered by Claude (Anthropic). No data leaves your machine
except the prompts you send to the Claude API.
"""

from __future__ import annotations

import argparse
import json
import os
import random
import re
import shutil
import sqlite3
import subprocess
import sys
import textwrap
from contextlib import contextmanager
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Any

from flask import Flask, g, jsonify, render_template, request

# Anthropic SDK is optional at import time so the app boots without a key.
try:
    from anthropic import Anthropic
except ImportError:  # pragma: no cover
    Anthropic = None  # type: ignore

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:  # pragma: no cover
    pass


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)
DB_PATH = DATA_DIR / "pipeline.db"

DEFAULT_MODEL = os.getenv("CLAUDE_MODEL", "claude-sonnet-4-5")
MAX_TOKENS = int(os.getenv("CLAUDE_MAX_TOKENS", "2048"))

app = Flask(__name__, static_folder="static", template_folder="templates")


# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------

SCHEMA = """
CREATE TABLE IF NOT EXISTS schema_version (
    version INTEGER PRIMARY KEY,
    applied_at TEXT
);

CREATE TABLE IF NOT EXISTS settings (
    key TEXT PRIMARY KEY,
    value TEXT
);

CREATE TABLE IF NOT EXISTS pillars (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    description TEXT,
    target_pct INTEGER DEFAULT 20,
    color TEXT DEFAULT '#6366f1',
    sort_order INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS ideas (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    pillar_id INTEGER REFERENCES pillars(id) ON DELETE SET NULL,
    title TEXT NOT NULL,
    hook TEXT,
    angle TEXT,
    source TEXT,
    status TEXT DEFAULT 'raw',  -- raw | drafted | discarded
    created_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS drafts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    idea_id INTEGER REFERENCES ideas(id) ON DELETE SET NULL,
    pillar_id INTEGER REFERENCES pillars(id) ON DELETE SET NULL,
    title TEXT,
    body TEXT NOT NULL,
    format TEXT DEFAULT 'story',  -- story | list | contrarian | tutorial | carousel | bts
    hook_score INTEGER,
    voice_score INTEGER,
    score_notes TEXT,
    status TEXT DEFAULT 'draft',  -- draft | ready | scheduled | published
    scheduled_for TEXT,
    posted_at TEXT,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS analytics (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    draft_id INTEGER REFERENCES drafts(id) ON DELETE CASCADE,
    impressions INTEGER DEFAULT 0,
    likes INTEGER DEFAULT 0,
    comments INTEGER DEFAULT 0,
    reposts INTEGER DEFAULT 0,
    follows INTEGER DEFAULT 0,
    profile_visits INTEGER DEFAULT 0,
    recorded_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS voice_samples (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    content TEXT NOT NULL,
    label TEXT,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS engagement_tasks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    draft_id INTEGER REFERENCES drafts(id) ON DELETE SET NULL,
    type TEXT DEFAULT 'comment',  -- comment | follow_up | respond
    details TEXT,
    due_date TEXT,
    completed INTEGER DEFAULT 0,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS reflections (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
    window_days INTEGER DEFAULT 7,
    summary TEXT,
    signals_json TEXT,
    ideas_created_json TEXT
);

CREATE TABLE IF NOT EXISTS topic_sources (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    url TEXT NOT NULL UNIQUE,
    kind TEXT DEFAULT 'rss',  -- rss | atom | web_search (future)
    enabled INTEGER DEFAULT 1,
    last_fetched_at TEXT,
    last_status TEXT,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS topics (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_id INTEGER REFERENCES topic_sources(id) ON DELETE CASCADE,
    external_id TEXT,
    url TEXT UNIQUE,
    title TEXT NOT NULL,
    summary TEXT,
    published_at TEXT,
    ingested_at TEXT DEFAULT CURRENT_TIMESTAMP,
    status TEXT DEFAULT 'new',  -- new | queued | used | dismissed
    pillar_id INTEGER REFERENCES pillars(id) ON DELETE SET NULL,
    idea_id INTEGER REFERENCES ideas(id) ON DELETE SET NULL
);
"""


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        g.db = conn
    return g.db


@app.teardown_appcontext
def close_db(_):
    db = g.pop("db", None)
    if db is not None:
        db.close()


@contextmanager
def db_cursor():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Schema migrations
#
# Each migration is a (version, function) pair. Functions take a connection,
# run their SQL, and must be idempotent (safe to re-run if interrupted).
# ---------------------------------------------------------------------------

CURRENT_SCHEMA_VERSION = 3


def _migration_v1(conn: sqlite3.Connection) -> None:
    """Baseline. SCHEMA covers all tables; nothing extra to do here."""
    pass


def _migration_v2(conn: sqlite3.Connection) -> None:
    """Add the reflections table for the weekly brain loop.

    Idempotent: SCHEMA already creates the table with IF NOT EXISTS, so this
    is a no-op on fresh installs. Existing DBs picked it up via the same
    executescript() in init_db() before MIGRATIONS run, so really we just need
    to be recorded as applied. Kept explicit for the schema_version trail.
    """
    conn.execute(
        "CREATE TABLE IF NOT EXISTS reflections ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT, "
        "created_at TEXT DEFAULT CURRENT_TIMESTAMP, "
        "window_days INTEGER DEFAULT 7, "
        "summary TEXT, "
        "signals_json TEXT, "
        "ideas_created_json TEXT)"
    )


def _migration_v3(conn: sqlite3.Connection) -> None:
    """Add topic_sources and topics tables for the RSS intake loop.

    Same pattern as v2: SCHEMA creates them via IF NOT EXISTS; this records
    the version. Kept explicit so ALTER-style migrations have a home later.
    """
    conn.execute(
        "CREATE TABLE IF NOT EXISTS topic_sources ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT, "
        "name TEXT NOT NULL, "
        "url TEXT NOT NULL UNIQUE, "
        "kind TEXT DEFAULT 'rss', "
        "enabled INTEGER DEFAULT 1, "
        "last_fetched_at TEXT, "
        "last_status TEXT, "
        "created_at TEXT DEFAULT CURRENT_TIMESTAMP)"
    )
    conn.execute(
        "CREATE TABLE IF NOT EXISTS topics ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT, "
        "source_id INTEGER REFERENCES topic_sources(id) ON DELETE CASCADE, "
        "external_id TEXT, "
        "url TEXT UNIQUE, "
        "title TEXT NOT NULL, "
        "summary TEXT, "
        "published_at TEXT, "
        "ingested_at TEXT DEFAULT CURRENT_TIMESTAMP, "
        "status TEXT DEFAULT 'new', "
        "pillar_id INTEGER REFERENCES pillars(id) ON DELETE SET NULL, "
        "idea_id INTEGER REFERENCES ideas(id) ON DELETE SET NULL)"
    )


MIGRATIONS: list[tuple[int, Any]] = [
    (1, _migration_v1),
    (2, _migration_v2),
    (3, _migration_v3),
]


def _applied_versions(conn: sqlite3.Connection) -> set[int]:
    rows = conn.execute("SELECT version FROM schema_version").fetchall()
    return {r["version"] for r in rows}


def init_db():
    with db_cursor() as conn:
        conn.executescript(SCHEMA)
        applied = _applied_versions(conn)
        for version, fn in MIGRATIONS:
            if version in applied:
                continue
            fn(conn)
            conn.execute(
                "INSERT INTO schema_version(version, applied_at) VALUES (?, ?)",
                (version, datetime.utcnow().isoformat(timespec="seconds")),
            )
        # Seed default pillars + voice samples + settings on first run only.
        # Each block is gated on its own table being empty so partial-seed
        # databases (e.g. user wiped pillars but kept voice) don't get re-seeded.
        cur = conn.execute("SELECT COUNT(*) AS n FROM pillars")
        if cur.fetchone()["n"] == 0:
            conn.executemany(
                "INSERT INTO pillars(name, description, target_pct, color, sort_order)"
                " VALUES (?, ?, ?, ?, ?)",
                DEFAULT_PILLARS,
            )
        cur = conn.execute("SELECT COUNT(*) AS n FROM voice_samples")
        if cur.fetchone()["n"] == 0 and DEFAULT_VOICE_SAMPLES:
            conn.executemany(
                "INSERT INTO voice_samples(content, label) VALUES (?, ?)",
                DEFAULT_VOICE_SAMPLES,
            )
        cur = conn.execute("SELECT COUNT(*) AS n FROM settings")
        if cur.fetchone()["n"] == 0:
            for k, v in DEFAULT_SETTINGS.items():
                conn.execute(
                    "INSERT INTO settings(key, value) VALUES (?, ?)", (k, v)
                )
        cur = conn.execute("SELECT COUNT(*) AS n FROM topic_sources")
        if cur.fetchone()["n"] == 0:
            conn.executemany(
                "INSERT INTO topic_sources(name, url, kind) VALUES (?, ?, ?)",
                DEFAULT_TOPIC_SOURCES,
            )


# ---------------------------------------------------------------------------
# Seed data — generic starter pillars and empty creator profile
#
# These constants only affect *fresh* installs (init_db inserts only when a
# table is empty). Existing databases are untouched. The first-run onboarding
# flow in the UI lets new users edit these in 30 seconds.
# ---------------------------------------------------------------------------

DEFAULT_PILLARS = [
    (
        "Industry insights",
        "Hot takes, trends, and observations from your domain. The 'I've been thinking about this' posts.",
        30, "#0ea5e9", 1,
    ),
    (
        "Tactical tutorials",
        "Step-by-step how-tos. Anything where the reader walks away able to do something they couldn't before.",
        25, "#10b981", 2,
    ),
    (
        "Personal stories",
        "Lessons learned the hard way. Career milestones. Mistakes that taught you something worth sharing.",
        20, "#f59e0b", 3,
    ),
    (
        "Tools and workflows",
        "What you use, why you use it, what you'd swap out. Reviews, comparisons, setups.",
        15, "#a855f7", 4,
    ),
    (
        "Behind the scenes",
        "The messy middle. Conference travel, side projects, the human stuff people actually relate to.",
        10, "#ef4444", 5,
    ),
]

# Empty by default. The onboarding modal asks new users to paste 3-5 of their
# best past posts so Claude has a real voice to imitate from day one.
DEFAULT_VOICE_SAMPLES: list[tuple[str, str]] = []

DEFAULT_SETTINGS = {
    "creator_name": "",
    "creator_handle": "",
    "creator_bio": "",
    "target_audience": "",
    "default_format": "story",
    "anthropic_api_key": "",
    "weekly_target": "5",
    "preferred_hours": "09:00,12:30,17:30",
}

# Default RSS feeds seeded on first run. Generic on purpose — the user adds,
# removes, or disables any of these from the Topics tab. Kept short so first
# fetch isn't slow. Only proven-working feeds; if you add a new one, manually
# curl it before committing.
DEFAULT_TOPIC_SOURCES = [
    ("Hacker News (front page)", "https://hnrss.org/frontpage", "rss"),
    ("dev.to (main feed)", "https://dev.to/feed", "rss"),
]


# ---------------------------------------------------------------------------
# AI helpers
# ---------------------------------------------------------------------------

def get_setting(key: str, default: str = "") -> str:
    with db_cursor() as conn:
        row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        return row["value"] if row else default


def set_setting(key: str, value: str):
    with db_cursor() as conn:
        conn.execute(
            "INSERT INTO settings(key, value) VALUES (?, ?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, value),
        )


# --- Backend selection: API key vs Claude Code CLI ---------------------------

def is_valid_api_key(k: str | None) -> bool:
    """Reject blanks, placeholders, and obvious junk so the API path doesn't
    spend 401s on values like 'sk-ant-...' that came from a template."""
    if not k:
        return False
    k = k.strip()
    if not k.startswith("sk-ant-"):
        return False
    # Real keys are >40 chars; placeholders are usually <20.
    if len(k) < 30:
        return False
    # Common placeholders contain literal "..." or look like obvious dummies.
    if "..." in k or k.lower().endswith(("-here", "-key")):
        return False
    return True


def auth_status() -> dict:
    """Detect which AI backend is available right now."""
    raw = os.getenv("ANTHROPIC_API_KEY") or get_setting("anthropic_api_key")
    api_key = raw if is_valid_api_key(raw) else ""
    cli_path = shutil.which("claude")
    cli_authed = False
    cli_version = ""
    if cli_path:
        try:
            r = subprocess.run(
                [cli_path, "--version"],
                capture_output=True, text=True, timeout=5,
            )
            cli_version = (r.stdout or "").strip()
            # CLI being on PATH is the strongest signal we have without making a
            # real call. We optimistically treat it as authed; if it's not, the
            # first generation will fail with a clear "claude login" message.
            cli_authed = bool(cli_version)
        except Exception:
            pass

    if api_key:
        provider = "api"
    elif cli_authed:
        provider = "cli"
    else:
        provider = "none"

    return {
        "provider": provider,
        "api_key_set": bool(api_key),
        "cli_installed": bool(cli_path),
        "cli_path": cli_path or "",
        "cli_version": cli_version,
    }


def _system_blocks_to_str(system: str | list) -> str:
    """Flatten a structured system prompt back to a plain string. Used for the
    CLI backend, which has no concept of prompt caching."""
    if isinstance(system, str):
        return system
    parts = []
    for block in system:
        if isinstance(block, str):
            parts.append(block)
        elif isinstance(block, dict):
            parts.append(block.get("text", ""))
    return "\n\n".join(p for p in parts if p)


def _call_via_api(system: str | list, user: str, model: str | None,
                  max_tokens: int | None) -> str:
    if Anthropic is None:
        raise RuntimeError(
            "anthropic SDK is not installed. Run: pip install anthropic"
        )
    raw = os.getenv("ANTHROPIC_API_KEY") or get_setting("anthropic_api_key")
    if not is_valid_api_key(raw):
        raise RuntimeError(
            "Anthropic API key looks invalid (placeholder or wrong format). "
            "Either save a real one in Settings or remove it to use Claude Code CLI."
        )
    client = Anthropic(api_key=raw)
    msg = client.messages.create(
        model=model or DEFAULT_MODEL,
        max_tokens=max_tokens or MAX_TOKENS,
        system=system,
        messages=[{"role": "user", "content": user}],
    )
    return "".join(block.text for block in msg.content if block.type == "text")


def _call_via_cli(system: str | list, user: str) -> str:
    """Use the local Claude Code CLI. Reuses the user's existing browser auth.

    Critical: we strip any ANTHROPIC_* env vars before invoking the CLI.
    Claude Code prefers an env-var key over its OAuth session, so if the
    parent process has a bad/placeholder key in its env (e.g. loaded from a
    poisoned .env), the CLI will fail too. Stripping forces it to fall back
    to the OAuth session from `claude login`.
    """
    cli_path = shutil.which("claude")
    if not cli_path:
        raise RuntimeError(
            "Claude Code CLI not found on PATH. Install it from "
            "https://docs.claude.com/en/docs/claude-code, then run `claude login`."
        )

    clean_env = {
        k: v for k, v in os.environ.items()
        if not k.startswith("ANTHROPIC_")
    }

    cmd = [
        cli_path,
        "-p", user,
        "--append-system-prompt", _system_blocks_to_str(system),
        "--output-format", "text",
    ]
    try:
        r = subprocess.run(
            cmd, capture_output=True, text=True, timeout=180, env=clean_env,
        )
    except FileNotFoundError as e:
        raise RuntimeError(f"Could not exec claude CLI: {e}")
    if r.returncode != 0:
        msg = (r.stderr or r.stdout or "").strip()
        # Common case: not authenticated yet
        if "auth" in msg.lower() or "login" in msg.lower() or "401" in msg:
            raise RuntimeError(
                "Claude Code CLI isn't authenticated. Run `claude login` "
                "in a terminal once, then come back."
            )
        raise RuntimeError(f"claude CLI error: {msg[:300]}")
    return r.stdout


def call_claude(system: str | list, user: str, model: str | None = None,
                max_tokens: int | None = None, json_mode: bool = False) -> str:
    """Run a single Claude completion via whichever backend is available.

    If the primary backend (API key) errors out with an auth/quota issue,
    we transparently fall back to the Claude Code CLI when it's available.
    """
    status = auth_status()
    text = None
    api_err: Exception | None = None

    if status["provider"] == "api":
        try:
            text = _call_via_api(system, user, model, max_tokens)
        except Exception as e:
            api_err = e
            if status["cli_installed"]:
                text = _call_via_cli(system, user)
            else:
                raise
    elif status["provider"] == "cli":
        text = _call_via_cli(system, user)
    else:
        raise RuntimeError(
            "No AI backend available. Either set an Anthropic API key in "
            "Settings, or install Claude Code and run `claude login` "
            "(https://docs.claude.com/en/docs/claude-code)."
        )

    if json_mode:
        text = text.strip()
        if text.startswith("```"):
            text = text.split("\n", 1)[-1]
            if text.endswith("```"):
                text = text.rsplit("```", 1)[0]
            if text.lstrip().lower().startswith("json"):
                text = text.split("\n", 1)[-1]
    return text.strip()


def voice_block() -> str:
    with db_cursor() as conn:
        rows = conn.execute(
            "SELECT content FROM voice_samples ORDER BY RANDOM() LIMIT 3"
        ).fetchall()
    if not rows:
        return ""
    samples = "\n\n---\n\n".join(r["content"] for r in rows)
    return f"\n\nVOICE SAMPLES (match this rhythm, sentence length, and energy):\n\n{samples}\n"


def winners_block(pillar_id: int | None = None, limit: int = 3) -> str:
    """Pull the top-performing published posts so the AI can see what's actually
    landing with the audience. Score = likes + 2 * comments (engagement-weighted).
    Restricted to a pillar when one is given so suggestions stay on-niche."""
    with db_cursor() as conn:
        sql = """
            SELECT d.body, d.format, p.name as pillar,
                   COALESCE(SUM(a.likes),0) as likes,
                   COALESCE(SUM(a.comments),0) as comments,
                   COALESCE(SUM(a.likes),0) + 2*COALESCE(SUM(a.comments),0) as score
            FROM drafts d
            JOIN analytics a ON a.draft_id = d.id
            LEFT JOIN pillars p ON p.id = d.pillar_id
            WHERE d.status = 'published'
        """
        params: list = []
        if pillar_id:
            sql += " AND d.pillar_id = ?"
            params.append(pillar_id)
        sql += " GROUP BY d.id HAVING score > 0 ORDER BY score DESC LIMIT ?"
        params.append(limit)
        rows = conn.execute(sql, tuple(params)).fetchall()

    if not rows:
        return ""
    parts = [
        "TOP PERFORMERS (these worked — same energy / hooks / structure win):"
    ]
    for r in rows:
        snippet = (r["body"] or "")[:600]
        parts.append(
            f"\n[format: {r['format']} · pillar: {r['pillar'] or 'unassigned'} · "
            f"{r['likes']} likes · {r['comments']} comments]\n{snippet}\n---"
        )
    return "\n".join(parts)


def discarded_block(limit: int = 20) -> str:
    """Recently-rejected ideas. The user said 'no' — don't recycle them,
    and don't pitch anything semantically close either. We include the hook
    line so the model has more than a title to compare against."""
    with db_cursor() as conn:
        rows = conn.execute(
            "SELECT title, hook FROM ideas WHERE status='discarded' "
            "ORDER BY created_at DESC LIMIT ?",
            (limit,),
        ).fetchall()
    if not rows:
        return ""
    lines = []
    for r in rows:
        line = f"- {r['title']}"
        if r["hook"]:
            line += f" — hook: {r['hook']}"
        lines.append(line)
    body = "\n".join(lines)
    return (
        "DISCARDED PATTERNS (the user already said no to these — don't pitch "
        "them again, and don't pitch anything SEMANTICALLY SIMILAR either, "
        "even with different wording or framing):\n" + body
    )


def memory_snapshot() -> dict:
    """What is the AI seeing right now?"""
    with db_cursor() as conn:
        winners = conn.execute(
            "SELECT d.id, d.title, d.format, p.name as pillar, "
            "COALESCE(SUM(a.likes),0)+2*COALESCE(SUM(a.comments),0) as score "
            "FROM drafts d JOIN analytics a ON a.draft_id=d.id "
            "LEFT JOIN pillars p ON p.id=d.pillar_id "
            "WHERE d.status='published' "
            "GROUP BY d.id HAVING score > 0 ORDER BY score DESC LIMIT 5"
        ).fetchall()
        voice_n = conn.execute(
            "SELECT COUNT(*) n FROM voice_samples"
        ).fetchone()["n"]
        discarded_n = conn.execute(
            "SELECT COUNT(*) n FROM ideas WHERE status='discarded'"
        ).fetchone()["n"]
        published_n = conn.execute(
            "SELECT COUNT(*) n FROM drafts WHERE status='published'"
        ).fetchone()["n"]
    return {
        "voice_samples": voice_n,
        "published": published_n,
        "discarded": discarded_n,
        "top_performers": [dict(r) for r in winners],
    }


def creator_block() -> str:
    bio = get_setting("creator_bio")
    audience = get_setting("target_audience")
    return textwrap.dedent(f"""
        CREATOR PROFILE:
        {bio}

        TARGET AUDIENCE:
        {audience}
    """).strip()


SYSTEM_BASE = textwrap.dedent("""
    You are a senior LinkedIn ghostwriter who has helped DevOps, cloud, and developer-advocate
    creators grow from zero to 100K+ followers. You write in the creator's voice — never
    yours. You understand what makes posts spread on LinkedIn in 2026: a compelling first
    line, short paragraphs, line breaks for breathing room, a clear point of view, and a
    soft CTA that invites conversation.

    Hard rules:
    • Never use em dashes (—). Use periods or line breaks instead.
    • Never use the words "delve", "leverage", "unlock", "unleash", or "game-changer".
    • Never start a post with "In today's fast-paced world" or any equivalent.
    • Avoid hashtags unless explicitly asked. If used, max 3, lowercase.
    • Each line in the post should be a deliberate beat, not filler.
    • The first line is everything. Make people stop scrolling.
""").strip()


# ---------------------------------------------------------------------------
# Routes — pages
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


# ---------------------------------------------------------------------------
# Routes — settings
# ---------------------------------------------------------------------------

@app.get("/api/settings")
def api_settings_get():
    with db_cursor() as conn:
        rows = conn.execute("SELECT key, value FROM settings").fetchall()
    settings = {r["key"]: r["value"] for r in rows}
    # Don't ship the actual API key to the frontend; just whether it's set.
    settings.pop("anthropic_api_key", "")
    settings["auth"] = auth_status()
    settings["anthropic_api_key_set"] = settings["auth"]["api_key_set"]
    return jsonify(settings)


@app.get("/api/auth/status")
def api_auth_status():
    return jsonify(auth_status())


@app.post("/api/auth/test")
def api_auth_test():
    """Probe each available backend so the user sees exactly what works."""
    status = auth_status()
    sys_prompt = "You are a health check. Reply with exactly: OK"
    user_prompt = "Say OK."

    used = None
    reply = None
    notes: list[str] = []

    # Try API key first if it's set
    if status["api_key_set"]:
        try:
            reply = _call_via_api(sys_prompt, user_prompt, None, 10)
            used = "api"
        except Exception as e:
            notes.append(f"API key failed: {e}")

    # Fall back to CLI
    if used is None and status["cli_installed"]:
        try:
            reply = _call_via_cli(sys_prompt, user_prompt)
            used = "cli"
        except Exception as e:
            notes.append(f"CLI failed: {e}")

    if used:
        return jsonify({
            "ok": True, "reply": (reply or "").strip(),
            "used": used, "notes": notes, "auth": status,
        })
    return jsonify({
        "ok": False,
        "error": " · ".join(notes) or "No AI backend available.",
        "auth": status,
    })


@app.post("/api/settings/clear-api-key")
def api_settings_clear_api_key():
    set_setting("anthropic_api_key", "")
    return jsonify({"ok": True})


@app.put("/api/settings")
def api_settings_put():
    payload = request.get_json(force=True)
    for k, v in payload.items():
        if k == "anthropic_api_key_set":
            continue
        set_setting(k, str(v))
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Routes — onboarding
#
# First-run flag is "no creator name set". The frontend pings status on boot;
# if first_run, it pops a modal asking for name/handle/bio/audience and
# optionally a few past posts to seed the voice library.
# ---------------------------------------------------------------------------

@app.get("/api/onboarding/status")
def api_onboarding_status():
    return jsonify({
        "first_run": not (get_setting("creator_name") or "").strip(),
        "has_voice_samples": _row_count("voice_samples") > 0,
        "has_pillars": _row_count("pillars") > 0,
    })


@app.post("/api/onboarding/complete")
def api_onboarding_complete():
    payload = request.get_json(force=True)
    profile_keys = ("creator_name", "creator_handle", "creator_bio",
                    "target_audience", "weekly_target", "default_format")
    for k in profile_keys:
        if k in payload:
            set_setting(k, str(payload[k] or ""))
    samples = payload.get("voice_samples") or []
    if samples:
        with db_cursor() as conn:
            for s in samples:
                content = (s.get("content") or "").strip() if isinstance(s, dict) else str(s).strip()
                if not content:
                    continue
                label = (s.get("label") or "seed").strip() if isinstance(s, dict) else "seed"
                conn.execute(
                    "INSERT INTO voice_samples(content, label) VALUES (?, ?)",
                    (content, label),
                )
    return jsonify({"ok": True})


def _row_count(table: str) -> int:
    with db_cursor() as conn:
        return conn.execute(f"SELECT COUNT(*) AS n FROM {table}").fetchone()["n"]


# ---------------------------------------------------------------------------
# Routes — backup / migrate
#
# Pure-data dump and restore. JSON, no binary blobs, schema_version included
# so a future restore knows whether the source was older or newer.
# ---------------------------------------------------------------------------

EXPORT_TABLES = (
    "settings", "pillars", "ideas", "drafts", "analytics",
    "voice_samples", "engagement_tasks", "reflections",
    "topic_sources", "topics",
)


def export_to_dict(redact_api_key: bool = True) -> dict:
    payload: dict[str, Any] = {
        "exported_at": datetime.utcnow().isoformat(timespec="seconds"),
        "schema_version": CURRENT_SCHEMA_VERSION,
        "tables": {},
    }
    with db_cursor() as conn:
        for tbl in EXPORT_TABLES:
            rows = conn.execute(f"SELECT * FROM {tbl}").fetchall()
            data = [dict(r) for r in rows]
            if redact_api_key and tbl == "settings":
                for r in data:
                    if r.get("key") == "anthropic_api_key":
                        r["value"] = ""
            payload["tables"][tbl] = data
    return payload


def import_from_dict(payload: dict, *, mode: str = "replace") -> dict:
    """Restore a backup. mode='replace' wipes existing user-data tables first;
    mode='merge' inserts on top of what's there (may produce duplicates).

    schema_version of the backup must be <= current; we don't downgrade."""
    backup_version = int(payload.get("schema_version") or 0)
    if backup_version > CURRENT_SCHEMA_VERSION:
        raise ValueError(
            f"Backup is from a newer schema (v{backup_version}); current app is "
            f"v{CURRENT_SCHEMA_VERSION}. Update the app first."
        )
    tables = payload.get("tables") or {}
    counts: dict[str, int] = {}
    with db_cursor() as conn:
        if mode == "replace":
            # Order matters because of FK refs; child tables first.
            for tbl in ("analytics", "engagement_tasks", "topics",
                        "topic_sources", "drafts", "ideas", "voice_samples",
                        "pillars", "settings", "reflections"):
                conn.execute(f"DELETE FROM {tbl}")
        for tbl in EXPORT_TABLES:
            rows = tables.get(tbl) or []
            counts[tbl] = 0
            for r in rows:
                cols = list(r.keys())
                placeholders = ",".join("?" for _ in cols)
                col_list = ",".join(cols)
                conn.execute(
                    f"INSERT INTO {tbl}({col_list}) VALUES ({placeholders})",
                    tuple(r[c] for c in cols),
                )
                counts[tbl] += 1
    return {"imported": counts, "from_schema": backup_version}


@app.get("/api/backup/export")
def api_backup_export():
    redact = request.args.get("redact", "1") != "0"
    payload = export_to_dict(redact_api_key=redact)
    fname = f"cadence-backup-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.json"
    return (
        json.dumps(payload, indent=2, ensure_ascii=False),
        200,
        {
            "Content-Type": "application/json; charset=utf-8",
            "Content-Disposition": f'attachment; filename="{fname}"',
        },
    )


@app.post("/api/backup/import")
def api_backup_import():
    """Body: {payload: {...backup json...}, mode: 'replace' | 'merge'}."""
    body = request.get_json(force=True)
    payload = body.get("payload")
    mode = body.get("mode") or "replace"
    if not isinstance(payload, dict):
        return jsonify({"error": "payload must be a JSON object"}), 400
    if mode not in ("replace", "merge"):
        return jsonify({"error": "mode must be 'replace' or 'merge'"}), 400
    try:
        result = import_from_dict(payload, mode=mode)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"ok": True, **result})


# ---------------------------------------------------------------------------
# Routes — pillars
# ---------------------------------------------------------------------------

@app.get("/api/pillars")
def api_pillars_get():
    with db_cursor() as conn:
        rows = conn.execute(
            "SELECT * FROM pillars ORDER BY sort_order, id"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/pillars")
def api_pillars_create():
    p = request.get_json(force=True)
    with db_cursor() as conn:
        cur = conn.execute(
            "INSERT INTO pillars(name, description, target_pct, color, sort_order) "
            "VALUES (?, ?, ?, ?, ?)",
            (
                p.get("name", "Untitled"),
                p.get("description", ""),
                int(p.get("target_pct", 20)),
                p.get("color", "#6366f1"),
                int(p.get("sort_order", 99)),
            ),
        )
        return jsonify({"id": cur.lastrowid})


@app.put("/api/pillars/<int:pid>")
def api_pillars_update(pid):
    p = request.get_json(force=True)
    fields = ["name", "description", "target_pct", "color", "sort_order"]
    sets = ", ".join(f"{f}=?" for f in fields if f in p)
    values = [p[f] for f in fields if f in p]
    if not sets:
        return jsonify({"ok": True})
    with db_cursor() as conn:
        conn.execute(f"UPDATE pillars SET {sets} WHERE id=?", (*values, pid))
    return jsonify({"ok": True})


@app.delete("/api/pillars/<int:pid>")
def api_pillars_delete(pid):
    with db_cursor() as conn:
        conn.execute("DELETE FROM pillars WHERE id=?", (pid,))
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Routes — ideas
# ---------------------------------------------------------------------------

@app.get("/api/ideas")
def api_ideas_list():
    status = request.args.get("status")
    with db_cursor() as conn:
        if status:
            rows = conn.execute(
                "SELECT i.*, p.name as pillar_name, p.color as pillar_color "
                "FROM ideas i LEFT JOIN pillars p ON p.id=i.pillar_id "
                "WHERE i.status=? ORDER BY i.created_at DESC",
                (status,),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT i.*, p.name as pillar_name, p.color as pillar_color "
                "FROM ideas i LEFT JOIN pillars p ON p.id=i.pillar_id "
                "ORDER BY i.created_at DESC"
            ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/ideas")
def api_ideas_create():
    p = request.get_json(force=True)
    with db_cursor() as conn:
        cur = conn.execute(
            "INSERT INTO ideas(pillar_id, title, hook, angle, source, status) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (
                p.get("pillar_id"),
                p.get("title", "Untitled"),
                p.get("hook", ""),
                p.get("angle", ""),
                p.get("source", "manual"),
                p.get("status", "raw"),
            ),
        )
        return jsonify({"id": cur.lastrowid})


@app.put("/api/ideas/<int:iid>")
def api_ideas_update(iid):
    p = request.get_json(force=True)
    fields = ["pillar_id", "title", "hook", "angle", "source", "status"]
    sets = ", ".join(f"{f}=?" for f in fields if f in p)
    values = [p[f] for f in fields if f in p]
    if not sets:
        return jsonify({"ok": True})
    with db_cursor() as conn:
        conn.execute(f"UPDATE ideas SET {sets} WHERE id=?", (*values, iid))
    return jsonify({"ok": True})


@app.delete("/api/ideas/<int:iid>")
def api_ideas_delete(iid):
    with db_cursor() as conn:
        conn.execute("DELETE FROM ideas WHERE id=?", (iid,))
    return jsonify({"ok": True})


@app.post("/api/ideas/generate")
def api_ideas_generate():
    """Generate fresh post ideas, optionally constrained to a pillar / theme."""
    p = request.get_json(force=True)
    pillar_id = p.get("pillar_id")
    theme = (p.get("theme") or "").strip()
    count = int(p.get("count", 5))

    with db_cursor() as conn:
        if pillar_id:
            pillar = conn.execute(
                "SELECT * FROM pillars WHERE id=?", (pillar_id,)
            ).fetchone()
        else:
            pillar = None
        recent = conn.execute(
            "SELECT title FROM ideas ORDER BY created_at DESC LIMIT 20"
        ).fetchall()

    pillar_text = (
        f"\nPILLAR FOCUS:\n{pillar['name']} — {pillar['description']}"
        if pillar else
        "\nNo specific pillar — pick whichever pillar fits each idea best."
    )
    theme_text = f"\nTHEME / TRIGGER: {theme}" if theme else ""
    avoid = "\n".join(f"- {r['title']}" for r in recent) or "(none yet)"

    # Memory injection — winners and discarded patterns
    winners = winners_block(pillar_id=pillar_id, limit=3)
    discarded = discarded_block()
    memory_parts = [b for b in (winners, discarded) if b]
    memory_text = "\n\n".join(memory_parts)

    system = SYSTEM_BASE + "\n\n" + creator_block()
    user = textwrap.dedent(f"""
        Generate {count} LinkedIn post IDEAS for this creator.

        {pillar_text}
        {theme_text}

        AVOID repeating these recent ideas (by topic, by hook, or by framing):
        {avoid}

        {memory_text}

        SEMANTIC DEDUP RULES (this is the rule the model usually breaks):
        - Treat any idea above as "claimed" not just by title but by THEME.
          "Moving off Kubernetes" and "We killed K8s" are the same idea.
        - If your draft idea is semantically close to something in the recent
          or DISCARDED list, throw it out and pick a different angle.
        - Variety target: 5 ideas should cover 5 distinct themes, not 5
          re-skins of one theme.

        For each idea give:
        - title: 6-10 word working title
        - hook: the actual first line of the post (must stop the scroll)
        - angle: 1-2 sentences on the unique POV / story / data point
        - format: one of [story, list, contrarian, tutorial, bts]

        Return ONLY valid JSON in this shape:
        {{
          "ideas": [
            {{"title": "...", "hook": "...", "angle": "...", "format": "story"}},
            ...
          ]
        }}
    """).strip()

    try:
        text = call_claude(system, user, json_mode=True)
        data = json.loads(text)
        created = []
        with db_cursor() as conn:
            for it in data.get("ideas", []):
                cur = conn.execute(
                    "INSERT INTO ideas(pillar_id, title, hook, angle, source) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (
                        pillar_id,
                        it.get("title", "Untitled")[:200],
                        it.get("hook", ""),
                        it.get("angle", "") + f"\n\n[format: {it.get('format', 'story')}]",
                        "ai",
                    ),
                )
                created.append(cur.lastrowid)
        return jsonify({"ok": True, "created": created, "count": len(created)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ---------------------------------------------------------------------------
# Routes — drafts
# ---------------------------------------------------------------------------

@app.get("/api/drafts")
def api_drafts_list():
    status = request.args.get("status")
    with db_cursor() as conn:
        sql = (
            "SELECT d.*, p.name as pillar_name, p.color as pillar_color "
            "FROM drafts d LEFT JOIN pillars p ON p.id=d.pillar_id "
        )
        if status:
            sql += "WHERE d.status=? ORDER BY d.updated_at DESC"
            rows = conn.execute(sql, (status,)).fetchall()
        else:
            sql += "ORDER BY d.updated_at DESC"
            rows = conn.execute(sql).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/drafts/<int:did>")
def api_drafts_get(did):
    with db_cursor() as conn:
        row = conn.execute(
            "SELECT d.*, p.name as pillar_name, p.color as pillar_color "
            "FROM drafts d LEFT JOIN pillars p ON p.id=d.pillar_id "
            "WHERE d.id=?",
            (did,),
        ).fetchone()
    return jsonify(dict(row) if row else None)


@app.post("/api/drafts")
def api_drafts_create():
    p = request.get_json(force=True)
    with db_cursor() as conn:
        cur = conn.execute(
            "INSERT INTO drafts(idea_id, pillar_id, title, body, format, status) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (
                p.get("idea_id"),
                p.get("pillar_id"),
                p.get("title", ""),
                p.get("body", ""),
                p.get("format", "story"),
                p.get("status", "draft"),
            ),
        )
        return jsonify({"id": cur.lastrowid})


@app.put("/api/drafts/<int:did>")
def api_drafts_update(did):
    p = request.get_json(force=True)
    p["updated_at"] = datetime.utcnow().isoformat(timespec="seconds")
    fields = [
        "idea_id", "pillar_id", "title", "body", "format",
        "hook_score", "voice_score", "score_notes", "status",
        "scheduled_for", "posted_at", "updated_at",
    ]
    sets = ", ".join(f"{f}=?" for f in fields if f in p)
    values = [p[f] for f in fields if f in p]
    if not sets:
        return jsonify({"ok": True})
    with db_cursor() as conn:
        conn.execute(f"UPDATE drafts SET {sets} WHERE id=?", (*values, did))
    return jsonify({"ok": True})


@app.delete("/api/drafts/<int:did>")
def api_drafts_delete(did):
    with db_cursor() as conn:
        conn.execute("DELETE FROM drafts WHERE id=?", (did,))
    return jsonify({"ok": True})


FORMAT_GUIDES = {
    "story": (
        "STORY format. Structure: hook line → setup (1-2 lines) → tension/conflict → "
        "turning point → lesson → soft CTA question. 120-220 words. Conversational."
    ),
    "list": (
        "LIST format. Hook line → 1 sentence framing → 5 to 7 numbered items, each 1-3 "
        "lines, parallel grammar → wrap-up line → CTA. Use newline between items."
    ),
    "contrarian": (
        "CONTRARIAN format. Bold first line stating a belief most people hold → "
        "'Here's why they're wrong' → your angle backed by a story or data → reframe → "
        "CTA invite to disagree."
    ),
    "tutorial": (
        "TUTORIAL format. Hook = the painful problem → 'Here's the 4-step fix:' → "
        "numbered steps with concrete commands or actions → result → CTA."
    ),
    "carousel": (
        "CAROUSEL script. Output 8-10 slides as 'Slide 1:', 'Slide 2:'... Slide 1 is a "
        "scroll-stopping hook. Slides 2-8 are one idea each. Slide 9 is summary. Slide "
        "10 is CTA. After the slides, give a 60-90 word LinkedIn caption."
    ),
    "bts": (
        "BEHIND-THE-SCENES format. Vulnerable hook (a moment of doubt, failure, or "
        "surprise) → context → what happened → what you learned → CTA inviting others "
        "to share."
    ),
}


@app.post("/api/drafts/generate")
def api_drafts_generate():
    """Turn an idea (or freeform brief) into a full draft."""
    p = request.get_json(force=True)
    idea_id = p.get("idea_id")
    fmt = p.get("format") or get_setting("default_format", "story")
    extra = (p.get("extra") or "").strip()

    idea = None
    with db_cursor() as conn:
        if idea_id:
            idea = conn.execute(
                "SELECT i.*, p.name as pillar_name FROM ideas i "
                "LEFT JOIN pillars p ON p.id=i.pillar_id WHERE i.id=?",
                (idea_id,),
            ).fetchone()

    brief = ""
    pillar_id = None
    if idea:
        pillar_id = idea["pillar_id"]
        brief = (
            f"TITLE: {idea['title']}\n"
            f"HOOK SEED: {idea['hook'] or '(write your own)'}\n"
            f"ANGLE: {idea['angle'] or '(figure it out)'}\n"
            f"PILLAR: {idea['pillar_name'] or 'general'}\n"
        )
    else:
        brief = p.get("brief", "Open brief. Surprise me.")
        pillar_id = p.get("pillar_id")

    winners = winners_block(pillar_id=pillar_id, limit=2)
    memory_text = f"\n\n{winners}" if winners else ""

    system = SYSTEM_BASE + "\n\n" + creator_block() + voice_block() + memory_text
    user = textwrap.dedent(f"""
        Write ONE LinkedIn post.

        {brief}

        {FORMAT_GUIDES.get(fmt, FORMAT_GUIDES['story'])}

        {('EXTRA INSTRUCTIONS: ' + extra) if extra else ''}

        Output ONLY the post body, ready to paste into LinkedIn. Use real line breaks
        (\\n) between paragraphs and items. No preamble, no commentary, no markdown.
    """).strip()

    try:
        body = call_claude(system, user)
        with db_cursor() as conn:
            cur = conn.execute(
                "INSERT INTO drafts(idea_id, pillar_id, title, body, format, status) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (
                    idea_id,
                    pillar_id,
                    (idea["title"] if idea else (p.get("title") or "Untitled")),
                    body,
                    fmt,
                    "draft",
                ),
            )
            draft_id = cur.lastrowid
            if idea_id:
                conn.execute(
                    "UPDATE ideas SET status='drafted' WHERE id=?", (idea_id,)
                )
        return jsonify({"ok": True, "id": draft_id, "body": body})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/api/drafts/<int:did>/rewrite")
def api_drafts_rewrite(did):
    p = request.get_json(force=True)
    instruction = p.get("instruction", "Make it tighter and punchier.")
    with db_cursor() as conn:
        row = conn.execute("SELECT * FROM drafts WHERE id=?", (did,)).fetchone()
    if not row:
        return jsonify({"ok": False, "error": "Draft not found"}), 404

    system = SYSTEM_BASE + "\n\n" + creator_block() + voice_block()
    user = textwrap.dedent(f"""
        Rewrite this LinkedIn post per the instruction below. Keep the writer's voice.

        ORIGINAL POST:
        ---
        {row['body']}
        ---

        INSTRUCTION:
        {instruction}

        Output ONLY the new post body. No preamble.
    """).strip()

    try:
        new_body = call_claude(system, user)
        with db_cursor() as conn:
            conn.execute(
                "UPDATE drafts SET body=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (new_body, did),
            )
        return jsonify({"ok": True, "body": new_body})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/api/drafts/<int:did>/repurpose")
def api_drafts_repurpose(did):
    """Take a published winner and spin out 3 fresh format variants."""
    with db_cursor() as conn:
        row = conn.execute(
            "SELECT d.*, p.name as pillar_name FROM drafts d "
            "LEFT JOIN pillars p ON p.id=d.pillar_id WHERE d.id=?",
            (did,),
        ).fetchone()
    if not row:
        return jsonify({"ok": False, "error": "Draft not found"}), 404

    system = SYSTEM_BASE + "\n\n" + creator_block() + voice_block()
    user = textwrap.dedent(f"""
        This LinkedIn post landed. Generate 3 NEW posts that ride the same
        winning idea but feel completely fresh — different format, different hook,
        different angle on the same insight. Don't paraphrase. Re-imagine.

        ORIGINAL POST:
        ---
        {row['body']}
        ---

        Each variant must use a DIFFERENT format from this set:
        story, list, contrarian, tutorial, bts.

        Return ONLY valid JSON:
        {{"variants": [
          {{"format": "list", "title": "6-10 word working title", "body": "the full post"}},
          {{"format": "...", "title": "...", "body": "..."}},
          {{"format": "...", "title": "...", "body": "..."}}
        ]}}
    """).strip()

    try:
        text = call_claude(system, user, json_mode=True, max_tokens=2400)
        data = json.loads(text)
        created = []
        with db_cursor() as conn:
            for v in data.get("variants", []):
                cur = conn.execute(
                    "INSERT INTO drafts(pillar_id, title, body, format, status) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (
                        row["pillar_id"],
                        (v.get("title") or "Repurpose")[:200],
                        v.get("body", ""),
                        v.get("format", "story"),
                        "draft",
                    ),
                )
                created.append(cur.lastrowid)
        return jsonify({"ok": True, "created": created, "count": len(created)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.get("/api/memory")
def api_memory():
    """What context the AI is currently drawing from."""
    return jsonify(memory_snapshot())


# ---------------------------------------------------------------------------
# Routes — brain (weekly reflection loop)
#
# A reflection reads the last N days of analytics, voice, and discarded ideas,
# asks Claude to summarise what's working and drop 3 fresh ideas into the
# pipeline. Triggered manually via the Dashboard button or `python app.py
# reflect`. No in-process scheduler — keeps Flask stateless.
# ---------------------------------------------------------------------------


def _gather_reflection_context(window_days: int) -> dict:
    """Pull the data the reflection prompt needs. Pure DB; no AI."""
    since = (datetime.utcnow() - timedelta(days=window_days)).isoformat(
        timespec="seconds"
    )
    with db_cursor() as conn:
        published = conn.execute(
            "SELECT d.id, d.title, d.format, p.name as pillar, "
            "COALESCE(SUM(a.impressions),0) impressions, "
            "COALESCE(SUM(a.likes),0) likes, "
            "COALESCE(SUM(a.comments),0) comments, "
            "COALESCE(SUM(a.follows),0) follows "
            "FROM drafts d "
            "LEFT JOIN analytics a ON a.draft_id = d.id "
            "LEFT JOIN pillars p ON p.id = d.pillar_id "
            "WHERE d.status='published' AND COALESCE(d.posted_at, d.updated_at) >= ? "
            "GROUP BY d.id ORDER BY likes DESC",
            (since,),
        ).fetchall()
        discarded = conn.execute(
            "SELECT title, hook FROM ideas WHERE status='discarded' "
            "AND created_at >= ? ORDER BY created_at DESC LIMIT 15",
            (since,),
        ).fetchall()
        recent_reflection = conn.execute(
            "SELECT summary FROM reflections ORDER BY created_at DESC LIMIT 1"
        ).fetchone()
        pillars = conn.execute(
            "SELECT name, target_pct FROM pillars ORDER BY sort_order"
        ).fetchall()
    return {
        "published": [dict(r) for r in published],
        "discarded": [dict(r) for r in discarded],
        "previous_reflection": (recent_reflection["summary"]
                                if recent_reflection else ""),
        "pillars": [dict(r) for r in pillars],
    }


def _format_reflection_table(rows: list[dict]) -> str:
    if not rows:
        return "(no posts published in this window)"
    lines = ["title | format | pillar | impressions | likes | comments | follows"]
    for r in rows:
        lines.append(
            f"{(r['title'] or '')[:60]} | {r['format']} | "
            f"{r['pillar'] or 'unassigned'} | {r['impressions']} | "
            f"{r['likes']} | {r['comments']} | {r['follows']}"
        )
    return "\n".join(lines)


def run_reflection(window_days: int = 7) -> dict:
    """Generate a reflection. Used by both the HTTP endpoint and the CLI."""
    ctx = _gather_reflection_context(window_days)
    table = _format_reflection_table(ctx["published"])
    discarded_lines = "\n".join(
        f"- {d['title']}" + (f" — hook: {d['hook']}" if d['hook'] else "")
        for d in ctx["discarded"]
    ) or "(none)"
    pillars_lines = ", ".join(
        f"{p['name']} ({p['target_pct']}%)" for p in ctx["pillars"]
    )
    prev = ctx["previous_reflection"] or "(no prior reflection)"

    # Two-block system prompt. The static part is identical across calls so
    # the API can cache it. The CLI fallback just joins everything back.
    static_system = (
        SYSTEM_BASE + "\n\n" + creator_block() + voice_block()
        + "\n\nYou are now acting as a weekly content coach. Be specific. "
          "Quote real numbers from the table when you can. Avoid generic "
          "advice. If a pillar is overperforming, name it."
    )
    system = [
        {"type": "text", "text": static_system,
         "cache_control": {"type": "ephemeral"}},
    ]

    user = textwrap.dedent(f"""
        Reflect on the last {window_days} days for this creator.

        ACTIVE PILLARS (with target mix %):
        {pillars_lines}

        PUBLISHED POSTS IN WINDOW:
        {table}

        RECENTLY DISCARDED IDEAS (don't pitch these or anything semantically
        similar):
        {discarded_lines}

        PREVIOUS REFLECTION (for continuity, may be stale):
        {prev}

        Return ONLY valid JSON in this exact shape:
        {{
          "summary": "5-7 sentence paragraph. Start with the single biggest signal. Name pillars and formats. End with one concrete experiment for next week.",
          "signals": {{
            "best_pillar": "name or null",
            "best_format": "story|list|contrarian|tutorial|carousel|bts or null",
            "weakest_pillar": "name or null",
            "topics_to_double_down_on": ["...", "...", "..."]
          }},
          "next_ideas": [
            {{"title": "6-10 word title", "hook": "the first line", "angle": "1-2 sentences", "format": "story|list|contrarian|tutorial|bts", "pillar": "name from active pillars or null"}},
            {{"title": "...", "hook": "...", "angle": "...", "format": "...", "pillar": "..."}},
            {{"title": "...", "hook": "...", "angle": "...", "format": "...", "pillar": "..."}}
          ]
        }}
    """).strip()

    text = call_claude(system, user, json_mode=True, max_tokens=1500)
    data = json.loads(text)

    # Map pillar names back to ids for the auto-created ideas
    with db_cursor() as conn:
        pillar_map = {
            r["name"].lower(): r["id"]
            for r in conn.execute("SELECT id, name FROM pillars").fetchall()
        }

    created_idea_ids: list[int] = []
    with db_cursor() as conn:
        for it in data.get("next_ideas", []) or []:
            pname = (it.get("pillar") or "").strip().lower()
            pid = pillar_map.get(pname)
            angle = (it.get("angle") or "")
            fmt = it.get("format") or "story"
            angle_with_format = f"{angle}\n\n[format: {fmt}]"
            cur = conn.execute(
                "INSERT INTO ideas(pillar_id, title, hook, angle, source) "
                "VALUES (?, ?, ?, ?, ?)",
                (
                    pid,
                    (it.get("title") or "Untitled")[:200],
                    it.get("hook", ""),
                    angle_with_format,
                    "auto-reflection",
                ),
            )
            created_idea_ids.append(cur.lastrowid)

        cur = conn.execute(
            "INSERT INTO reflections(window_days, summary, signals_json, "
            "ideas_created_json) VALUES (?, ?, ?, ?)",
            (
                window_days,
                data.get("summary", ""),
                json.dumps(data.get("signals") or {}),
                json.dumps(created_idea_ids),
            ),
        )
        reflection_id = cur.lastrowid

    return {
        "id": reflection_id,
        "summary": data.get("summary", ""),
        "signals": data.get("signals") or {},
        "ideas_created": created_idea_ids,
        "window_days": window_days,
    }


@app.post("/api/brain/reflect")
def api_brain_reflect():
    payload = request.get_json(silent=True) or {}
    try:
        window = int(payload.get("window_days", 7))
    except (TypeError, ValueError):
        window = 7
    window = max(1, min(window, 90))
    try:
        result = run_reflection(window_days=window)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.get("/api/brain/reflections")
def api_brain_reflections():
    try:
        limit = int(request.args.get("limit", "10"))
    except ValueError:
        limit = 10
    limit = max(1, min(limit, 50))
    with db_cursor() as conn:
        rows = conn.execute(
            "SELECT id, created_at, window_days, summary, signals_json, "
            "ideas_created_json FROM reflections "
            "ORDER BY created_at DESC LIMIT ?",
            (limit,),
        ).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        try:
            d["signals"] = json.loads(d.pop("signals_json") or "{}")
        except json.JSONDecodeError:
            d["signals"] = {}
        try:
            d["ideas_created"] = json.loads(d.pop("ideas_created_json") or "[]")
        except json.JSONDecodeError:
            d["ideas_created"] = []
        out.append(d)
    return jsonify(out)


@app.post("/api/drafts/<int:did>/score")
def api_drafts_score(did):
    with db_cursor() as conn:
        row = conn.execute("SELECT * FROM drafts WHERE id=?", (did,)).fetchone()
    if not row:
        return jsonify({"ok": False, "error": "Draft not found"}), 404

    system = SYSTEM_BASE + "\n\n" + creator_block() + voice_block()
    user = textwrap.dedent(f"""
        Score this LinkedIn post on TWO dimensions, 1-10 integer each.

        1. HOOK STRENGTH (does the first line stop the scroll?)
           - 10 = unmissable, specific, emotional, contrarian or curiosity-driven
           - 5  = generic but not bad
           - 1  = puts the reader to sleep

        2. VOICE MATCH (does it sound like the creator's samples above?)
           - 10 = indistinguishable from the samples
           - 5  = on-topic but bland / generic LinkedIn voice
           - 1  = sounds AI-generated

        Then give 3 short, concrete improvement notes (one line each).

        POST:
        ---
        {row['body']}
        ---

        Return ONLY valid JSON:
        {{"hook_score": 0, "voice_score": 0, "notes": ["...","...","..."]}}
    """).strip()

    try:
        text = call_claude(system, user, json_mode=True, max_tokens=600)
        data = json.loads(text)
        notes = " • ".join(data.get("notes", []))
        with db_cursor() as conn:
            conn.execute(
                "UPDATE drafts SET hook_score=?, voice_score=?, score_notes=? "
                "WHERE id=?",
                (
                    int(data.get("hook_score", 0)),
                    int(data.get("voice_score", 0)),
                    notes,
                    did,
                ),
            )
        return jsonify({"ok": True, **data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ---------------------------------------------------------------------------
# Routes — topics (RSS / Atom intake)
#
# The "what should I post about?" feed. Pulls headlines from user-configured
# sources (HN, dev.to, vendor blogs), dedups by URL, and lets the user turn
# any topic into an angled idea via Claude. Triggered manually from the
# Topics tab — no in-process cron, mirroring the brain-loop philosophy.
# ---------------------------------------------------------------------------

TOPIC_FETCH_TIMEOUT_SEC = 10
TOPIC_FETCH_MAX_BYTES = 2 * 1024 * 1024  # 2 MB per feed; LinkedIn-sized
TOPIC_FETCH_MAX_ITEMS_PER_SOURCE = 30  # don't drown the table on first fetch
TOPIC_USER_AGENT = "Cadence/1.0 (+https://github.com/nomadicmehul/cadence)"


def _http_get_bounded(url: str, *, timeout: int, max_bytes: int) -> bytes:
    """Plain urllib GET with a hard byte ceiling. No redirects-handling magic
    beyond urllib's default (which handles up to ~10 hops automatically)."""
    import urllib.request
    req = urllib.request.Request(url, headers={"User-Agent": TOPIC_USER_AGENT})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        # iter_content equivalent: cap bytes ourselves so a hostile feed can't
        # OOM the process.
        data = bytearray()
        while True:
            chunk = resp.read(64 * 1024)
            if not chunk:
                break
            data.extend(chunk)
            if len(data) > max_bytes:
                raise ValueError(
                    f"Response exceeded {max_bytes} bytes; aborted"
                )
        return bytes(data)


def _looks_like_html(raw: bytes) -> bool:
    """Cheap content sniff. Strips leading whitespace, lowercases the first
    400 bytes, and checks for an HTML doctype or root tag. Good enough to
    decide whether to trigger feed autodiscovery."""
    head = raw[:400].lstrip().lower()
    return head.startswith(b"<!doctype html") or head.startswith(b"<html")


def _discover_feed_url(html_bytes: bytes, base_url: str) -> str | None:
    """Scan HTML bytes for an RSS / Atom autodiscovery <link rel=alternate>.

    Returns the absolutized feed URL if one is found, else None. Atom is
    preferred over RSS when both are present (slightly more standard, but
    either path works with feedparser downstream).
    """
    from html.parser import HTMLParser
    from urllib.parse import urljoin

    try:
        text = html_bytes.decode("utf-8", errors="replace")
    except Exception:
        return None

    found: list[tuple[str, str]] = []  # (type, href)

    class _Finder(HTMLParser):
        def handle_starttag(self, tag, attrs):
            if tag != "link":
                return
            ad = {k.lower(): (v or "") for k, v in attrs}
            rel = ad.get("rel", "").lower()
            t = ad.get("type", "").lower()
            href = ad.get("href")
            if not href or "alternate" not in rel:
                return
            if "rss" in t or "atom" in t:
                found.append((t, href))

        def error(self, message):  # never raise on malformed HTML
            pass

    parser = _Finder(convert_charrefs=True)
    try:
        # HTMLParser stops at </head> in pathological docs but most pages put
        # discovery links inside <head>, so a single feed() call is fine.
        parser.feed(text)
    except Exception:
        # HTMLParser can raise on truly broken input; swallow and treat as
        # "no feed found" rather than crashing the whole fetch.
        pass

    if not found:
        return None

    # Atom first, then RSS, then anything else.
    found.sort(key=lambda x: 0 if "atom" in x[0] else 1)
    return urljoin(base_url, found[0][1])


def _parse_feed_entries(raw: bytes) -> list[dict]:
    """Run feedparser on raw bytes; normalise the bits we actually care about."""
    try:
        import feedparser
    except ImportError:
        raise RuntimeError("feedparser not installed. Run: pip install feedparser")
    parsed = feedparser.parse(raw)
    out: list[dict] = []
    for entry in parsed.entries[:TOPIC_FETCH_MAX_ITEMS_PER_SOURCE]:
        link = (entry.get("link") or "").strip()
        title = (entry.get("title") or "").strip()
        if not link or not title:
            continue
        summary = (
            entry.get("summary") or entry.get("description") or ""
        ).strip()
        # feedparser parses dates into struct_time on `published_parsed`.
        published_iso = ""
        pp = entry.get("published_parsed") or entry.get("updated_parsed")
        if pp:
            try:
                published_iso = datetime(*pp[:6]).isoformat(timespec="seconds")
            except (TypeError, ValueError):
                published_iso = ""
        out.append({
            "external_id": (entry.get("id") or link)[:300],
            "url": link[:1000],
            "title": title[:500],
            "summary": summary[:2000],
            "published_at": published_iso,
        })
    return out


def fetch_topics(*, source_ids: list[int] | None = None) -> dict:
    """Pull from every enabled source (or a subset by id), dedup by URL,
    insert new rows. Per-source failures don't abort the batch."""
    with db_cursor() as conn:
        if source_ids:
            qmarks = ",".join("?" for _ in source_ids)
            sources = conn.execute(
                f"SELECT * FROM topic_sources WHERE id IN ({qmarks}) "
                "AND enabled=1",
                tuple(source_ids),
            ).fetchall()
        else:
            sources = conn.execute(
                "SELECT * FROM topic_sources WHERE enabled=1"
            ).fetchall()

    per_source: list[dict] = []
    inserted_total = 0
    for src in sources:
        result = {
            "source_id": src["id"], "source_name": src["name"],
            "inserted": 0, "skipped_existing": 0, "error": None,
            "discovered_url": None,
        }
        try:
            raw = _http_get_bounded(
                src["url"],
                timeout=TOPIC_FETCH_TIMEOUT_SEC,
                max_bytes=TOPIC_FETCH_MAX_BYTES,
            )

            # Autodiscovery: if the user pasted a homepage (e.g.
            # https://dev.to/t/googlecloud) rather than a feed URL, look in
            # the HTML for <link rel="alternate" type="application/rss+xml">
            # and refetch from there. One extra round-trip on misconfigured
            # sources, zero overhead on feeds that respond with XML directly.
            if _looks_like_html(raw):
                discovered = _discover_feed_url(raw, src["url"])
                if discovered and discovered != src["url"]:
                    result["discovered_url"] = discovered
                    raw = _http_get_bounded(
                        discovered,
                        timeout=TOPIC_FETCH_TIMEOUT_SEC,
                        max_bytes=TOPIC_FETCH_MAX_BYTES,
                    )

            entries = _parse_feed_entries(raw)
        except Exception as e:
            result["error"] = str(e)[:200]
            with db_cursor() as conn:
                conn.execute(
                    "UPDATE topic_sources SET last_fetched_at=?, last_status=? "
                    "WHERE id=?",
                    (datetime.utcnow().isoformat(timespec="seconds"),
                     f"error: {result['error']}", src["id"]),
                )
            per_source.append(result)
            continue

        with db_cursor() as conn:
            for e in entries:
                # UNIQUE on url means duplicate inserts raise; count and skip.
                try:
                    conn.execute(
                        "INSERT INTO topics(source_id, external_id, url, "
                        "title, summary, published_at) "
                        "VALUES (?, ?, ?, ?, ?, ?)",
                        (src["id"], e["external_id"], e["url"], e["title"],
                         e["summary"], e["published_at"]),
                    )
                    result["inserted"] += 1
                except sqlite3.IntegrityError:
                    result["skipped_existing"] += 1

            # Build an honest status message:
            #  - if entries is empty AND we just fetched HTML with no
            #    discoverable feed, tell the user the URL probably isn't a
            #    feed rather than masquerading as "ok: 0 new".
            #  - if we autodiscovered, mention the resolved URL so the user
            #    can update their source manually if they want.
            if entries:
                status = f"ok: {result['inserted']} new"
                if result["discovered_url"]:
                    status += f" (via {result['discovered_url']})"
            elif result["discovered_url"]:
                status = (
                    f"ok: 0 entries at autodiscovered "
                    f"{result['discovered_url']}"
                )
            elif _looks_like_html(raw):
                status = "ok: 0 entries (URL returned HTML with no RSS/Atom link)"
            else:
                status = "ok: 0 entries (feed parsed but had no items)"

            conn.execute(
                "UPDATE topic_sources SET last_fetched_at=?, last_status=? "
                "WHERE id=?",
                (datetime.utcnow().isoformat(timespec="seconds"),
                 status, src["id"]),
            )
        inserted_total += result["inserted"]
        per_source.append(result)

    return {
        "inserted_total": inserted_total,
        "sources_checked": len(sources),
        "per_source": per_source,
    }


@app.get("/api/topics/sources")
def api_topic_sources_list():
    with db_cursor() as conn:
        rows = conn.execute(
            "SELECT * FROM topic_sources ORDER BY id"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/topics/sources")
def api_topic_sources_create():
    p = request.get_json(force=True)
    name = (p.get("name") or "").strip()
    url = (p.get("url") or "").strip()
    if not name or not url:
        return jsonify({"error": "name and url are required"}), 400
    if not (url.startswith("http://") or url.startswith("https://")):
        return jsonify({"error": "url must be http(s)"}), 400
    try:
        with db_cursor() as conn:
            cur = conn.execute(
                "INSERT INTO topic_sources(name, url, kind, enabled) "
                "VALUES (?, ?, ?, ?)",
                (name, url, p.get("kind", "rss"),
                 1 if p.get("enabled", True) else 0),
            )
        return jsonify({"id": cur.lastrowid})
    except sqlite3.IntegrityError:
        return jsonify({"error": "that url is already a source"}), 409


@app.put("/api/topics/sources/<int:sid>")
def api_topic_sources_update(sid):
    p = request.get_json(force=True)
    fields = ["name", "url", "kind", "enabled"]
    sets = ", ".join(f"{f}=?" for f in fields if f in p)
    values = [p[f] for f in fields if f in p]
    if not sets:
        return jsonify({"ok": True})
    try:
        with db_cursor() as conn:
            conn.execute(
                f"UPDATE topic_sources SET {sets} WHERE id=?",
                (*values, sid),
            )
    except sqlite3.IntegrityError:
        return jsonify({"error": "that url is already a source"}), 409
    return jsonify({"ok": True})


@app.delete("/api/topics/sources/<int:sid>")
def api_topic_sources_delete(sid):
    with db_cursor() as conn:
        conn.execute("DELETE FROM topic_sources WHERE id=?", (sid,))
    return jsonify({"ok": True})


@app.post("/api/topics/fetch")
def api_topics_fetch():
    p = request.get_json(silent=True) or {}
    source_ids = p.get("source_ids") or None
    if source_ids is not None:
        try:
            source_ids = [int(s) for s in source_ids]
        except (TypeError, ValueError):
            return jsonify({"error": "source_ids must be a list of ints"}), 400
    try:
        result = fetch_topics(source_ids=source_ids)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.get("/api/topics")
def api_topics_list():
    status = request.args.get("status")
    source_id = request.args.get("source_id")
    limit = request.args.get("limit", "100")
    try:
        limit = max(1, min(int(limit), 500))
    except ValueError:
        limit = 100
    sql = (
        "SELECT t.*, s.name as source_name, p.name as pillar_name, "
        "p.color as pillar_color "
        "FROM topics t LEFT JOIN topic_sources s ON s.id = t.source_id "
        "LEFT JOIN pillars p ON p.id = t.pillar_id WHERE 1=1"
    )
    params: list = []
    if status:
        sql += " AND t.status=?"
        params.append(status)
    if source_id:
        try:
            params.append(int(source_id))
            sql += " AND t.source_id=?"
        except ValueError:
            pass
    sql += " ORDER BY COALESCE(t.published_at, t.ingested_at) DESC LIMIT ?"
    params.append(limit)
    with db_cursor() as conn:
        rows = conn.execute(sql, tuple(params)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.put("/api/topics/<int:tid>")
def api_topics_update(tid):
    p = request.get_json(force=True)
    fields = ["status", "pillar_id"]
    sets = ", ".join(f"{f}=?" for f in fields if f in p)
    values = [p[f] for f in fields if f in p]
    if not sets:
        return jsonify({"ok": True})
    with db_cursor() as conn:
        conn.execute(
            f"UPDATE topics SET {sets} WHERE id=?", (*values, tid)
        )
    return jsonify({"ok": True})


@app.delete("/api/topics/<int:tid>")
def api_topics_delete(tid):
    with db_cursor() as conn:
        conn.execute("DELETE FROM topics WHERE id=?", (tid,))
    return jsonify({"ok": True})


@app.post("/api/topics/<int:tid>/draft")
def api_topics_draft(tid):
    """Turn a topic into an angled idea. Calls Claude with the creator's
    voice + pillars so the resulting idea isn't a generic 'react to news'."""
    p = request.get_json(silent=True) or {}
    pillar_id = p.get("pillar_id")
    with db_cursor() as conn:
        topic = conn.execute(
            "SELECT t.*, s.name as source_name FROM topics t "
            "LEFT JOIN topic_sources s ON s.id = t.source_id "
            "WHERE t.id=?",
            (tid,),
        ).fetchone()
        if not topic:
            return jsonify({"ok": False, "error": "Topic not found"}), 404
        pillars = conn.execute(
            "SELECT id, name, description FROM pillars ORDER BY sort_order"
        ).fetchall()

    pillar_lines = "\n".join(
        f"- {pr['name']}: {pr['description']}" for pr in pillars
    ) or "(no pillars defined yet)"

    static_system = (
        SYSTEM_BASE + "\n\n" + creator_block() + voice_block()
        + "\n\nYou are now generating ONE post idea inspired by an external "
          "topic. The idea must reflect the creator's voice and pillars, not "
          "merely summarise the source. Aim for a personal angle, lesson, or "
          "contrarian take."
    )
    system = [
        {"type": "text", "text": static_system,
         "cache_control": {"type": "ephemeral"}},
    ]

    user = textwrap.dedent(f"""
        EXTERNAL TOPIC:
        Source: {topic['source_name'] or '(unknown)'}
        Title: {topic['title']}
        URL: {topic['url'] or ''}
        Summary: {topic['summary'] or ''}

        ACTIVE PILLARS:
        {pillar_lines}

        Return ONLY valid JSON in this exact shape:
        {{
          "title": "6-10 word working title for the post",
          "hook": "the actual first line of the post (must stop the scroll)",
          "angle": "1-2 sentences on the unique POV / story / data point this creator brings",
          "format": "story|list|contrarian|tutorial|bts",
          "pillar": "name from the active pillars above, or null"
        }}
    """).strip()

    try:
        text = call_claude(system, user, json_mode=True, max_tokens=600)
        data = json.loads(text)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    # Map pillar by case-insensitive name unless caller overrode
    pillar_match = None
    if pillar_id:
        try:
            pillar_match = int(pillar_id)
        except (TypeError, ValueError):
            pillar_match = None
    if pillar_match is None:
        name = (data.get("pillar") or "").strip().lower()
        for pr in pillars:
            if pr["name"].lower() == name:
                pillar_match = pr["id"]
                break

    fmt = data.get("format") or "story"
    angle_with_format = (data.get("angle") or "") + f"\n\n[format: {fmt}]"
    angle_with_source = (
        angle_with_format + f"\n\nSource: {topic['url'] or topic['source_name']}"
    )

    with db_cursor() as conn:
        cur = conn.execute(
            "INSERT INTO ideas(pillar_id, title, hook, angle, source) "
            "VALUES (?, ?, ?, ?, ?)",
            (
                pillar_match,
                (data.get("title") or topic["title"])[:200],
                data.get("hook", ""),
                angle_with_source,
                "topic-intake",
            ),
        )
        idea_id = cur.lastrowid
        conn.execute(
            "UPDATE topics SET status='used', idea_id=?, pillar_id=? "
            "WHERE id=?",
            (idea_id, pillar_match, tid),
        )

    return jsonify({
        "ok": True,
        "idea_id": idea_id,
        "topic_id": tid,
        "pillar_id": pillar_match,
        "data": data,
    })


# ---------------------------------------------------------------------------
# Routes — calendar / scheduling
# ---------------------------------------------------------------------------

@app.get("/api/calendar")
def api_calendar():
    start = request.args.get("start")
    end = request.args.get("end")
    with db_cursor() as conn:
        if start and end:
            rows = conn.execute(
                "SELECT d.*, p.name as pillar_name, p.color as pillar_color "
                "FROM drafts d LEFT JOIN pillars p ON p.id=d.pillar_id "
                "WHERE d.scheduled_for IS NOT NULL "
                "AND d.scheduled_for BETWEEN ? AND ? "
                "ORDER BY d.scheduled_for ASC",
                (start, end),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT d.*, p.name as pillar_name, p.color as pillar_color "
                "FROM drafts d LEFT JOIN pillars p ON p.id=d.pillar_id "
                "WHERE d.scheduled_for IS NOT NULL "
                "ORDER BY d.scheduled_for ASC"
            ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/calendar/auto-schedule")
def api_calendar_auto_schedule():
    """
    Pick READY drafts and schedule them across the next N days, respecting
    pillar percentages and preferred posting hours.
    """
    p = request.get_json(force=True)
    days = int(p.get("days", 14))
    target_per_week = int(get_setting("weekly_target", "5") or "5")
    hours = [
        h.strip() for h in (get_setting("preferred_hours", "09:00,17:30")).split(",")
        if h.strip()
    ] or ["09:00"]

    with db_cursor() as conn:
        ready = conn.execute(
            "SELECT * FROM drafts WHERE status IN ('draft','ready') "
            "AND scheduled_for IS NULL ORDER BY updated_at DESC"
        ).fetchall()
        pillars = {
            r["id"]: r for r in
            conn.execute("SELECT * FROM pillars").fetchall()
        }

    if not ready:
        return jsonify({"ok": True, "scheduled": 0, "message": "No drafts ready."})

    # Build slots: every ~ (7/target_per_week) days at preferred hours
    slot_dates = []
    today = date.today()
    interval = max(1, round(7 / max(1, target_per_week)))
    cur = today
    while cur <= today + timedelta(days=days):
        if cur.weekday() < 5:  # weekdays only
            slot_dates.append(cur)
        cur += timedelta(days=interval)

    slots: list[datetime] = []
    for d in slot_dates:
        for h in hours:
            try:
                hh, mm = h.split(":")
                slots.append(datetime(d.year, d.month, d.day, int(hh), int(mm)))
            except ValueError:
                continue
    slots = slots[: len(ready)]

    # Order drafts to spread pillars
    by_pillar: dict[int | None, list] = {}
    for d in ready:
        by_pillar.setdefault(d["pillar_id"], []).append(d)
    interleaved = []
    while any(by_pillar.values()):
        for pid in list(by_pillar.keys()):
            if by_pillar[pid]:
                interleaved.append(by_pillar[pid].pop(0))

    scheduled = 0
    with db_cursor() as conn:
        for draft, slot in zip(interleaved, slots):
            conn.execute(
                "UPDATE drafts SET status='scheduled', scheduled_for=?, "
                "updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (slot.isoformat(timespec="minutes"), draft["id"]),
            )
            scheduled += 1
    return jsonify({"ok": True, "scheduled": scheduled})


# ---------------------------------------------------------------------------
# Routes — analytics
# ---------------------------------------------------------------------------

@app.get("/api/analytics")
def api_analytics_list():
    with db_cursor() as conn:
        rows = conn.execute(
            "SELECT a.*, d.title as draft_title, d.body as draft_body, "
            "d.posted_at as posted_at, p.name as pillar_name, p.color as pillar_color "
            "FROM analytics a "
            "JOIN drafts d ON d.id = a.draft_id "
            "LEFT JOIN pillars p ON p.id = d.pillar_id "
            "ORDER BY a.recorded_at DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/analytics")
def api_analytics_create():
    p = request.get_json(force=True)
    with db_cursor() as conn:
        cur = conn.execute(
            "INSERT INTO analytics(draft_id, impressions, likes, comments, "
            "reposts, follows, profile_visits) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (
                int(p["draft_id"]),
                int(p.get("impressions", 0) or 0),
                int(p.get("likes", 0) or 0),
                int(p.get("comments", 0) or 0),
                int(p.get("reposts", 0) or 0),
                int(p.get("follows", 0) or 0),
                int(p.get("profile_visits", 0) or 0),
            ),
        )
        # mark draft published if not already
        conn.execute(
            "UPDATE drafts SET status='published', "
            "posted_at = COALESCE(posted_at, CURRENT_TIMESTAMP) WHERE id=?",
            (int(p["draft_id"]),),
        )
        return jsonify({"id": cur.lastrowid})


def _xlsx_read_all_sheets(xlsx_bytes: bytes) -> list[dict]:
    """Read every sheet in the workbook, score each by post-analytics signal,
    return them sorted best-first. Each entry has: name, score, rows (csv text),
    row_count, preview (first 5 non-blank rows as list-of-list)."""
    import csv as _csv
    import io
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    # IMPORTANT: don't use read_only=True. LinkedIn's xlsx exports omit or
    # misreport the <dimension> tag, which read-only mode trusts blindly,
    # causing openpyxl to see only the top-left cell of each sheet.
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    keywords = (
        "impression", "view", "like", "reaction", "comment",
        "repost", "share", "engagement", "click", "post url", "post text",
        "post title", "follower", "created date", "published",
    )

    def cell_str(v) -> str:
        if v is None:
            return ""
        if hasattr(v, "isoformat"):
            return v.isoformat()
        return str(v)

    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[list[str]] = []
        # Iterate by explicit max_row/max_column to be robust against dimension
        # weirdness. ws.max_row/max_column reflect actual cell content even if
        # the dimension XML lies.
        max_r = ws.max_row or 0
        max_c = ws.max_column or 0
        for r in range(1, max_r + 1):
            cells = [cell_str(ws.cell(row=r, column=c).value)
                     for c in range(1, max_c + 1)]
            if any(c.strip() for c in cells):
                rows.append(cells)
        if not rows:
            continue
        buf = io.StringIO()
        w = _csv.writer(buf)
        for r in rows:
            w.writerow(r)
        sheet_csv = buf.getvalue()
        low = sheet_csv.lower()
        score = sum(low.count(k) for k in keywords)
        nm = name.lower()
        if "post" in nm or "top" in nm:
            score += 50
        if "discovery" in nm or "engagement" in nm:
            score += 10
        if "follower" in nm or "demographic" in nm or "summary" in nm:
            score -= 5
        if len(rows) < 2:
            score -= 30
        sheets.append({
            "name": name,
            "score": score,
            "csv": sheet_csv,
            "row_count": len(rows),
            "preview": rows[:5],
        })

    sheets.sort(key=lambda s: s["score"], reverse=True)
    return sheets


@app.post("/api/analytics/import-preview")
def api_analytics_import_preview():
    """
    Parse a LinkedIn analytics export (CSV or XLSX) and auto-match each row
    to an existing draft by body-snippet or date. The frontend shows the
    result and lets the user confirm or override.
    """
    import csv as _csv
    import io
    import base64
    payload = request.get_json(force=True)
    csv_text = payload.get("csv", "")
    xlsx_b64 = payload.get("xlsx", "")

    chosen_sheet = ""
    sheets_available: list[str] = []
    sheet_override = (payload.get("sheet_override") or "").strip()
    sheets_meta: list[dict] = []

    if xlsx_b64:
        try:
            xlsx_bytes = base64.b64decode(xlsx_b64)
        except Exception as e:
            return jsonify({"ok": False, "error": f"Bad base64 xlsx: {e}"}), 400
        try:
            sheets_meta = _xlsx_read_all_sheets(xlsx_bytes)
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 400
        if not sheets_meta:
            return jsonify({"ok": False, "error": "No non-empty sheets in workbook"}), 400
        sheets_available = [s["name"] for s in sheets_meta]

        # If user picked a specific sheet, honor that. Otherwise use highest score.
        chosen = None
        if sheet_override:
            for s in sheets_meta:
                if s["name"] == sheet_override:
                    chosen = s
                    break
        if chosen is None:
            chosen = sheets_meta[0]
        csv_text = chosen["csv"]
        chosen_sheet = chosen["name"]

    if not csv_text or not csv_text.strip():
        return jsonify({"ok": False, "error": "Empty file"}), 400

    # Some LinkedIn exports prefix the actual table with a title row, a blank
    # line, a generated-on timestamp, etc. Look for the first row that smells
    # like a header (multiple columns, contains an analytics keyword OR contains
    # a "post"/"date"/"url" word). If nothing matches, just take the first
    # non-blank row as the header.
    raw_reader = _csv.reader(io.StringIO(csv_text))
    all_rows = [r for r in raw_reader if any((c or "").strip() for c in r)]

    HEADER_KEYWORDS = (
        "impression", "view", "like", "reaction", "comment",
        "repost", "share", "engagement", "click",
        "post url", "post text", "post title", "post link",
        "created date", "published", "date posted", "posted on",
    )
    header_idx = -1
    for i, r in enumerate(all_rows):
        joined = ",".join((c or "").lower() for c in r)
        if len(r) >= 2 and any(k in joined for k in HEADER_KEYWORDS):
            header_idx = i
            break
    if header_idx == -1:
        # No header keyword found. Take the first multi-column row as header.
        for i, r in enumerate(all_rows):
            if len(r) >= 2:
                header_idx = i
                break
    if header_idx == -1:
        # Truly nothing to work with. If this came from xlsx, show all sheets
        # so the user can pick a different one manually.
        previews = [
            {"name": s["name"], "row_count": s["row_count"],
             "score": s["score"], "preview": s["preview"]}
            for s in sheets_meta
        ] if sheets_meta else []
        return jsonify({
            "ok": False,
            "error": "Could not find a usable header row.",
            "sheet_used": chosen_sheet,
            "sheets_preview": previews,
        }), 400

    rows = all_rows[header_idx:]
    if len(rows) < 2:
        previews = [
            {"name": s["name"], "row_count": s["row_count"],
             "score": s["score"], "preview": s["preview"]}
            for s in sheets_meta
        ] if sheets_meta else []
        return jsonify({
            "ok": False,
            "error": (f"Sheet '{chosen_sheet}' has a header but no data rows."
                      if chosen_sheet else
                      "Need a header and at least one data row."),
            "sheet_used": chosen_sheet,
            "sheets_preview": previews,
        }), 400

    headers_raw = rows[0]
    headers = [(h or "").strip().lower() for h in headers_raw]

    # LinkedIn's "Top posts" sheet sometimes contains TWO side-by-side
    # subtables: e.g. cols A-C = "Post URL · Date · Engagements" (top by
    # engagement), cols E-G = "Post URL · Date · Impressions" (top by
    # impressions). Same posts, different rankings, separated by a blank
    # column. We detect this by counting URL-like headers; if there's more
    # than one, we parse each subtable independently and merge by URL.
    url_idx_positions = [
        i for i, h in enumerate(headers)
        if ("url" in h or "link" in h) and "engagement" not in h
    ]
    subtable_ranges: list[tuple[int, int]] = []
    if len(url_idx_positions) >= 2:
        for k, start in enumerate(url_idx_positions):
            end = (url_idx_positions[k + 1]
                   if k + 1 < len(url_idx_positions)
                   else len(headers))
            subtable_ranges.append((start, end))
    else:
        subtable_ranges.append((0, len(headers)))

    def find_col_in(slice_headers: list[str], *candidates: str) -> int:
        """Find the first column index (relative to the full row) whose
        lowercased header contains any candidate keyword. Returns -1 on miss."""
        offset, sub = slice_headers[0], slice_headers[1]
        for cand in candidates:
            for j, h in enumerate(sub):
                if cand in h:
                    return offset + j
        return -1

    def detect_cols(start: int, end: int) -> dict:
        """Detect column indices within a [start, end) header slice. Returns
        the same shape `col` used to live in, but anchored to the subtable."""
        sub_headers = headers[start:end]
        sh = (start, sub_headers)
        # 'engagements' is LinkedIn's aggregate column (likes+comments+reposts
        # combined). When the export omits the breakdown, we map it to likes
        # so the value isn't silently lost. The /preview response flags this
        # so the UI can show a hint.
        likes_col = find_col_in(sh, "reaction", "like")
        engagement_col = find_col_in(sh, "engagement")
        used_aggregate = False
        if likes_col == -1 and engagement_col != -1:
            likes_col = engagement_col
            used_aggregate = True
        return {
            "date":     find_col_in(sh, "date", "created", "posted", "publish"),
            "impr":     find_col_in(sh, "impression", "view"),
            "likes":    likes_col,
            "comments": find_col_in(sh, "comment"),
            "reposts":  find_col_in(sh, "repost", "share"),
            "follows":  find_col_in(sh, "follow"),
            "url":      find_col_in(sh, "url", "link"),
            "text":     find_col_in(sh, "post text", "post title", "caption",
                                    "message", "content", "body"),
            "_used_engagement_aggregate": used_aggregate,
        }

    subtable_cols = [detect_cols(s, e) for (s, e) in subtable_ranges]

    # The first subtable's columns are the "primary" detected shape we expose
    # to the UI. Aggregate-engagement flag wins if any subtable used it.
    col = dict(subtable_cols[0])
    used_engagement_aggregate = any(
        c.get("_used_engagement_aggregate") for c in subtable_cols
    )

    with db_cursor() as conn:
        drafts = conn.execute(
            "SELECT id, body, posted_at, scheduled_for, title, status "
            "FROM drafts ORDER BY id DESC"
        ).fetchall()

    def to_int(s: str) -> int:
        s = (s or "").strip().replace(",", "")
        if not s:
            return 0
        # tolerate "1.2K" / "1,234" / floats
        try:
            if s.lower().endswith("k"):
                return int(float(s[:-1]) * 1000)
            return int(float(s))
        except ValueError:
            return 0

    def parse_iso(s: str):
        if not s:
            return None
        s = s.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%b %d, %Y"):
            try:
                return datetime.strptime(s[:25].split("T")[0], fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(s[:25])
        except ValueError:
            return None

    # Parse each subtable separately. Each subtable contributes a partial
    # record per row; we merge across subtables keyed by URL afterward so the
    # engagement count from subtable 1 and the impressions from subtable 2
    # land on the same record (rather than on whatever post happens to share
    # a row index between the two rankings).
    def g_of(row, idx: int) -> str:
        return (row[idx] or "").strip() if 0 <= idx < len(row) else ""

    METRIC_KEYS = ("impressions", "likes", "comments", "reposts", "follows")

    by_url: dict[str, dict] = {}
    ordered_urls: list[str] = []
    unkeyed: list[dict] = []  # rows that had no URL — emit at the end

    for sub_col in subtable_cols:
        for row in rows[1:]:
            url = g_of(row, sub_col["url"])
            date = g_of(row, sub_col["date"])
            snippet = g_of(row, sub_col["text"])[:120]
            contrib = {
                "impressions": to_int(g_of(row, sub_col["impr"])),
                "likes":       to_int(g_of(row, sub_col["likes"])),
                "comments":    to_int(g_of(row, sub_col["comments"])),
                "reposts":     to_int(g_of(row, sub_col["reposts"])),
                "follows":     to_int(g_of(row, sub_col["follows"])),
            }
            if not url and not any(contrib.values()) and not date:
                continue  # empty row, ignore
            if not url:
                # Unkeyable row from this subtable; we can't merge it safely.
                # Keep it standalone so the user sees something rather than
                # losing the data.
                unkeyed.append({
                    "date": date, "snippet": snippet, "url": "",
                    **contrib,
                })
                continue
            existing = by_url.get(url)
            if existing is None:
                by_url[url] = {
                    "url": url, "date": date, "snippet": snippet,
                    **contrib,
                }
                ordered_urls.append(url)
            else:
                # Merge: take whichever value is non-zero per metric, prefer
                # newer (the subtable being processed). Existing date/snippet
                # are kept unless missing.
                if not existing.get("date") and date:
                    existing["date"] = date
                if not existing.get("snippet") and snippet:
                    existing["snippet"] = snippet
                for k in METRIC_KEYS:
                    if not existing.get(k) and contrib.get(k):
                        existing[k] = contrib[k]

    parsed: list[dict] = []
    for i, url in enumerate(ordered_urls):
        rec = by_url[url]
        parsed.append({
            "row_idx": i,
            "date": rec["date"],
            "snippet": rec["snippet"],
            "url": rec["url"],
            "impressions": rec["impressions"],
            "likes":       rec["likes"],
            "comments":    rec["comments"],
            "reposts":     rec["reposts"],
            "follows":     rec["follows"],
            "matched_draft_id": None,
            "match_reason": "",
        })
    for j, rec in enumerate(unkeyed):
        parsed.append({
            "row_idx": len(parsed) + j,
            "date": rec["date"],
            "snippet": rec["snippet"],
            "url": "",
            "impressions": rec["impressions"],
            "likes":       rec["likes"],
            "comments":    rec["comments"],
            "reposts":     rec["reposts"],
            "follows":     rec["follows"],
            "matched_draft_id": None,
            "match_reason": "",
        })

    # Pass 1: snippet match (high confidence). Each draft can only claim one row.
    claimed: set[int] = set()
    for item in parsed:
        snip = (item["snippet"] or "").lower()[:35]
        if not snip:
            continue
        for d in drafts:
            if d["id"] in claimed:
                continue
            body = (d["body"] or "").lower()
            if snip and snip in body:
                item["matched_draft_id"] = d["id"]
                item["match_reason"] = "snippet"
                claimed.add(d["id"])
                break

    # Pass 2: date proximity for still-unmatched rows, against still-unclaimed drafts.
    for item in parsed:
        if item["matched_draft_id"]:
            continue
        post_date = parse_iso(item["date"])
        if not post_date:
            continue
        best = None
        best_delta = 99999
        for d in drafts:
            if d["id"] in claimed:
                continue
            for fld in ("posted_at", "scheduled_for"):
                v = d[fld]
                if not v:
                    continue
                dd = parse_iso(v)
                if not dd:
                    continue
                delta = abs((post_date - dd).days)
                if delta <= 3 and delta < best_delta:
                    best, best_delta = d["id"], delta
        if best:
            item["matched_draft_id"] = best
            item["match_reason"] = f"date(±{best_delta}d)"
            claimed.add(best)

    sheets_preview = [
        {"name": s["name"], "row_count": s["row_count"],
         "score": s["score"], "preview": s["preview"]}
        for s in sheets_meta
    ] if sheets_meta else []

    notes: list[str] = []
    if used_engagement_aggregate:
        notes.append(
            "LinkedIn's export didn't break engagement into likes/comments/"
            "reposts. We mapped the aggregate 'Engagements' column to likes "
            "so the value isn't lost. You can edit individual rows after "
            "import to redistribute."
        )
    if len(subtable_ranges) > 1:
        notes.append(
            f"Detected {len(subtable_ranges)} side-by-side subtables in the "
            "chosen sheet. Posts were merged by URL across them so the "
            "impressions / engagement counts land on the same row."
        )

    return jsonify({
        "ok": True,
        "headers": headers_raw,
        "detected": col,
        "parsed": parsed,
        "sheet_used": chosen_sheet,
        "sheets_available": sheets_available,
        "sheets_preview": sheets_preview,
        "notes": notes,
        "used_engagement_aggregate": used_engagement_aggregate,
        "subtable_count": len(subtable_ranges),
        "drafts": [
            {"id": d["id"], "title": d["title"] or "",
             "preview": (d["body"] or "")[:80], "status": d["status"]}
            for d in drafts
        ],
    })


def _title_from_import_row(r: dict) -> str:
    """Best-effort title for an auto-created historical draft. Order of
    preference: an explicit title on the row, the snippet (first 60 chars),
    or 'Imported post · <date>'. Activity-ID URLs don't carry post text
    so we have no body to mine."""
    explicit = (r.get("title") or "").strip()
    if explicit:
        return explicit[:200]
    snippet = (r.get("snippet") or "").strip()
    if snippet:
        return snippet[:60]
    date = (r.get("date") or "").strip()
    if date:
        return f"Imported post · {date}"
    return "Imported post"


def _parse_iso_for_posted_at(s: str) -> str | None:
    """Same date parser as in /import-preview, returns ISO string or None."""
    if not s:
        return None
    s = s.strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%b %d, %Y"):
        try:
            return datetime.strptime(s[:25].split("T")[0], fmt).isoformat(
                timespec="seconds"
            )
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s[:25]).isoformat(timespec="seconds")
    except ValueError:
        return None


CREATE_NEW_SENTINEL = "__new__"


@app.post("/api/analytics/import-commit")
def api_analytics_import_commit():
    """Insert the user-confirmed rows.

    draft_id semantics:
      - integer / numeric string  -> attach analytics to that existing draft
      - "__new__"                  -> create a fresh draft and attach
      - falsy / missing            -> skip this row (the user picked "skip")

    Historical-bulk-import workflow: when the user is on-boarding their
    existing LinkedIn analytics, none of the rows have an in-Cadence draft
    to match against. The "__new__" path creates a published-status draft
    using the URL + date so the analytics row has a home and future
    winners_block() queries can use the engagement signal.
    """
    payload = request.get_json(force=True)
    rows = payload.get("rows", [])
    created = 0
    skipped = 0
    new_drafts: list[int] = []
    with db_cursor() as conn:
        for r in rows:
            did_raw = r.get("draft_id")
            did: int | None = None
            if did_raw == CREATE_NEW_SENTINEL:
                # Create the placeholder draft, then attach the row.
                # Body stays empty because LinkedIn URN URLs don't include
                # the post text. The user can paste body text later for
                # any post they want to mine for voice training.
                posted_at = (_parse_iso_for_posted_at(r.get("date") or "")
                             or datetime.utcnow().isoformat(timespec="seconds"))
                cur = conn.execute(
                    "INSERT INTO drafts(idea_id, pillar_id, title, body, "
                    "format, status, posted_at) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (
                        None, None,
                        _title_from_import_row(r),
                        (r.get("url") or "")[:1000],  # URL goes in body as
                                                      # a backreference
                        "story",
                        "published",
                        posted_at,
                    ),
                )
                did = cur.lastrowid
                new_drafts.append(did)
            elif did_raw:
                try:
                    did = int(did_raw)
                except (TypeError, ValueError):
                    skipped += 1
                    continue
            else:
                skipped += 1
                continue

            conn.execute(
                "INSERT INTO analytics(draft_id, impressions, likes, comments, "
                "reposts, follows) VALUES (?, ?, ?, ?, ?, ?)",
                (
                    did,
                    int(r.get("impressions", 0) or 0),
                    int(r.get("likes", 0) or 0),
                    int(r.get("comments", 0) or 0),
                    int(r.get("reposts", 0) or 0),
                    int(r.get("follows", 0) or 0),
                ),
            )
            # Only flip status / set posted_at for drafts we didn't just
            # create (new ones already have status='published').
            if did_raw != CREATE_NEW_SENTINEL:
                conn.execute(
                    "UPDATE drafts SET status='published', "
                    "posted_at = COALESCE(posted_at, CURRENT_TIMESTAMP) "
                    "WHERE id=?",
                    (did,),
                )
            created += 1
    return jsonify({
        "ok": True,
        "created": created,
        "skipped": skipped,
        "new_drafts": new_drafts,
    })


@app.get("/api/analytics/insights")
def api_analytics_insights():
    """Roll up analytics + ask Claude what's working."""
    with db_cursor() as conn:
        rows = conn.execute(
            "SELECT d.id, d.title, d.format, p.name as pillar, "
            "SUM(a.impressions) impressions, SUM(a.likes) likes, "
            "SUM(a.comments) comments, SUM(a.follows) follows "
            "FROM analytics a JOIN drafts d ON d.id=a.draft_id "
            "LEFT JOIN pillars p ON p.id=d.pillar_id "
            "GROUP BY d.id ORDER BY likes DESC LIMIT 30"
        ).fetchall()

    if not rows:
        return jsonify({"insight": "Log some analytics first, then I'll spot patterns."})

    table = "title | format | pillar | impr | likes | comments | follows\n"
    for r in rows:
        table += (
            f"{(r['title'] or '')[:60]} | {r['format']} | {r['pillar']} | "
            f"{r['impressions']} | {r['likes']} | {r['comments']} | {r['follows']}\n"
        )

    system = SYSTEM_BASE + "\n\n" + creator_block()
    user = textwrap.dedent(f"""
        Here is the creator's recent post performance.

        {table}

        Give 5 SHORT, specific, actionable insights:
        - Which pillar is overperforming and why
        - Which format is winning
        - Which topics to do MORE of
        - What pattern in their hooks works best
        - One concrete experiment to run next week

        Use plain prose, max 6 lines total. No bullet markers.
    """).strip()

    try:
        text = call_claude(system, user, max_tokens=600)
        return jsonify({"insight": text})
    except Exception as e:
        return jsonify({"insight": f"Could not run insights: {e}"})


# ---------------------------------------------------------------------------
# Routes — engagement
# ---------------------------------------------------------------------------

@app.get("/api/engagement")
def api_engagement_list():
    with db_cursor() as conn:
        rows = conn.execute(
            "SELECT * FROM engagement_tasks ORDER BY completed ASC, due_date ASC, id DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/engagement")
def api_engagement_create():
    p = request.get_json(force=True)
    with db_cursor() as conn:
        cur = conn.execute(
            "INSERT INTO engagement_tasks(draft_id, type, details, due_date) "
            "VALUES (?, ?, ?, ?)",
            (
                p.get("draft_id"),
                p.get("type", "comment"),
                p.get("details", ""),
                p.get("due_date"),
            ),
        )
        return jsonify({"id": cur.lastrowid})


@app.put("/api/engagement/<int:eid>")
def api_engagement_update(eid):
    p = request.get_json(force=True)
    fields = ["type", "details", "due_date", "completed"]
    sets = ", ".join(f"{f}=?" for f in fields if f in p)
    values = [p[f] for f in fields if f in p]
    if not sets:
        return jsonify({"ok": True})
    with db_cursor() as conn:
        conn.execute(f"UPDATE engagement_tasks SET {sets} WHERE id=?", (*values, eid))
    return jsonify({"ok": True})


@app.delete("/api/engagement/<int:eid>")
def api_engagement_delete(eid):
    with db_cursor() as conn:
        conn.execute("DELETE FROM engagement_tasks WHERE id=?", (eid,))
    return jsonify({"ok": True})


_URL_ONLY_RE = re.compile(r"^https?://\S+$")


@app.post("/api/engagement/comments")
def api_engagement_comments():
    """Generate 3 thoughtful comment templates for a target post.

    Requires the BODY TEXT of the post, not a URL — Claude has no way to
    fetch URLs from here. We reject URL-only input with a clear 400 so the
    spinner doesn't hang while Claude hallucinates against a slug.
    """
    p = request.get_json(force=True)
    target = (p.get("target_post") or "").strip()
    if not target:
        return jsonify({
            "ok": False,
            "error": "Paste the body text of the LinkedIn post (not a URL).",
        }), 400
    if _URL_ONLY_RE.match(target):
        return jsonify({
            "ok": False,
            "error": ("That looks like just a URL. Cadence can't fetch "
                      "LinkedIn URLs. Open the post on LinkedIn, copy the "
                      "post text, and paste that instead."),
        }), 400

    # Two-block system: static portion (with voice samples!) marked cacheable.
    # voice_block was previously omitted here, which silently bypassed the
    # voice-consistency rule. Comments now actually sound like the creator.
    static_system = (
        SYSTEM_BASE + "\n\n" + creator_block() + voice_block()
        + "\n\nYou are writing comments the creator will leave on someone "
          "else's LinkedIn post. Comments must read in the creator's voice "
          "as shown in the samples above. No 'great post!' sycophancy. "
          "Each comment must add a specific story, data point, or counterpoint."
    )
    system = [
        {"type": "text", "text": static_system,
         "cache_control": {"type": "ephemeral"}},
    ]

    user = textwrap.dedent(f"""
        Write 3 short LinkedIn comments to leave on this post. Each comment must:
        - Add value (specific story, example, or counterpoint)
        - Sound like the creator (match the voice samples above)
        - Be 2-4 lines max
        - Not be sycophantic ("great post!" / "amazing!" — banned)
        - Not summarise the post back at the author

        TARGET POST (this is the body text the user wants to comment on):
        ---
        {target}
        ---

        Return ONLY JSON: {{"comments": ["...","...","..."]}}
    """).strip()
    try:
        text = call_claude(system, user, json_mode=True, max_tokens=600)
        return jsonify({"ok": True, **json.loads(text)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ---------------------------------------------------------------------------
# Routes — voice samples
# ---------------------------------------------------------------------------

@app.get("/api/voice")
def api_voice_list():
    with db_cursor() as conn:
        rows = conn.execute(
            "SELECT * FROM voice_samples ORDER BY id DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/voice")
def api_voice_create():
    p = request.get_json(force=True)
    with db_cursor() as conn:
        cur = conn.execute(
            "INSERT INTO voice_samples(content, label) VALUES (?, ?)",
            (p.get("content", ""), p.get("label", "")),
        )
        return jsonify({"id": cur.lastrowid})


@app.delete("/api/voice/<int:vid>")
def api_voice_delete(vid):
    with db_cursor() as conn:
        conn.execute("DELETE FROM voice_samples WHERE id=?", (vid,))
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Routes — dashboard
# ---------------------------------------------------------------------------

@app.get("/api/dashboard")
def api_dashboard():
    with db_cursor() as conn:
        ideas_raw = conn.execute(
            "SELECT COUNT(*) n FROM ideas WHERE status='raw'"
        ).fetchone()["n"]
        drafts_n = conn.execute(
            "SELECT COUNT(*) n FROM drafts WHERE status IN ('draft','ready')"
        ).fetchone()["n"]
        scheduled_n = conn.execute(
            "SELECT COUNT(*) n FROM drafts WHERE status='scheduled'"
        ).fetchone()["n"]
        published_n = conn.execute(
            "SELECT COUNT(*) n FROM drafts WHERE status='published'"
        ).fetchone()["n"]
        next_up = conn.execute(
            "SELECT d.id, d.title, d.scheduled_for, p.name as pillar_name, p.color as pillar_color "
            "FROM drafts d LEFT JOIN pillars p ON p.id=d.pillar_id "
            "WHERE d.status='scheduled' AND d.scheduled_for >= datetime('now') "
            "ORDER BY d.scheduled_for ASC LIMIT 5"
        ).fetchall()
        totals = conn.execute(
            "SELECT COALESCE(SUM(impressions),0) i, COALESCE(SUM(likes),0) l, "
            "COALESCE(SUM(comments),0) c, COALESCE(SUM(follows),0) f "
            "FROM analytics"
        ).fetchone()
        pillar_mix = conn.execute(
            "SELECT p.name, p.color, COUNT(d.id) n "
            "FROM pillars p LEFT JOIN drafts d ON d.pillar_id=p.id "
            "AND d.status='published' GROUP BY p.id ORDER BY p.sort_order"
        ).fetchall()
    return jsonify({
        "counts": {
            "ideas_raw": ideas_raw,
            "drafts": drafts_n,
            "scheduled": scheduled_n,
            "published": published_n,
        },
        "next_up": [dict(r) for r in next_up],
        "totals": dict(totals),
        "pillar_mix": [dict(r) for r in pillar_mix],
    })


# ---------------------------------------------------------------------------
# Boot
# ---------------------------------------------------------------------------

init_db()


def _cli_export(path: str, *, include_api_key: bool) -> int:
    payload = export_to_dict(redact_api_key=not include_api_key)
    Path(path).write_text(
        json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    totals = {t: len(payload["tables"][t]) for t in EXPORT_TABLES}
    print(f"Exported {sum(totals.values())} rows to {path}")
    for t, n in totals.items():
        print(f"  {t}: {n}")
    if include_api_key:
        print("WARNING: --include-api-key was set. Treat this file as a secret.")
    return 0


def _cli_reflect(window_days: int) -> int:
    try:
        result = run_reflection(window_days=window_days)
    except Exception as e:
        print(f"Reflection failed: {e}", file=sys.stderr)
        return 2
    print(f"Reflection #{result['id']} (window={result['window_days']}d)")
    print("-" * 60)
    print(result["summary"])
    print("-" * 60)
    if result["signals"]:
        print("Signals:")
        for k, v in result["signals"].items():
            print(f"  {k}: {v}")
    n = len(result["ideas_created"])
    print(f"Dropped {n} fresh idea(s) into the pipeline.")
    return 0


def _cli_import(path: str, *, mode: str, yes: bool) -> int:
    p = Path(path)
    if not p.exists():
        print(f"File not found: {path}", file=sys.stderr)
        return 2
    payload = json.loads(p.read_text(encoding="utf-8"))
    if mode == "replace" and not yes:
        print("This will DELETE all current data and replace it with the backup.")
        confirm = input("Type 'yes' to continue: ").strip().lower()
        if confirm != "yes":
            print("Aborted.")
            return 1
    try:
        result = import_from_dict(payload, mode=mode)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 2
    total = sum(result["imported"].values())
    print(f"Imported {total} rows from {path} (mode={mode}).")
    for t, n in result["imported"].items():
        print(f"  {t}: {n}")
    return 0


def _build_cli() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Cadence — run the server, or back up / restore your data.",
    )
    sub = parser.add_subparsers(dest="cmd")

    p_export = sub.add_parser("export", help="Dump all data to a JSON file.")
    p_export.add_argument("path", help="Output JSON path (e.g. backup.json)")
    p_export.add_argument(
        "--include-api-key", action="store_true",
        help="Include the saved Anthropic API key in the dump (default: redacted).",
    )

    p_import = sub.add_parser("import", help="Restore from a JSON backup.")
    p_import.add_argument("path", help="Backup JSON path")
    p_import.add_argument(
        "--mode", choices=("replace", "merge"), default="replace",
        help="replace = wipe current data first; merge = insert on top.",
    )
    p_import.add_argument(
        "-y", "--yes", action="store_true", help="Skip the confirmation prompt.",
    )

    p_reflect = sub.add_parser(
        "reflect",
        help="Run the weekly brain reflection — same as the Dashboard button.",
    )
    p_reflect.add_argument(
        "--days", type=int, default=7,
        help="Lookback window in days (default: 7).",
    )

    sub.add_parser("serve", help="Start the web server (default).")
    return parser


if __name__ == "__main__":
    parser = _build_cli()
    args = parser.parse_args()
    if args.cmd == "export":
        sys.exit(_cli_export(args.path, include_api_key=args.include_api_key))
    if args.cmd == "import":
        sys.exit(_cli_import(args.path, mode=args.mode, yes=args.yes))
    if args.cmd == "reflect":
        sys.exit(_cli_reflect(window_days=max(1, min(args.days, 90))))
    # Default: serve
    port = int(os.getenv("PORT", "5050"))
    print(f"\n  Cadence running at http://127.0.0.1:{port}\n")
    app.run(host="127.0.0.1", port=port, debug=True)
