"""Tests for the LinkedIn analytics xlsx importer, especially the cases that
broke against a real export:

  1. The "Top posts" sheet contains TWO side-by-side subtables (e.g. cols
     A-C = top by Engagements, cols E-G = top by Impressions). The parser
     needs to detect this and merge by URL — without that, the impressions
     and engagement values land on different posts.
  2. LinkedIn often only exposes an aggregate "Engagements" column, not the
     likes/comments/reposts breakdown. We map the aggregate into `likes` so
     the value isn't lost, and surface a note to the UI.
  3. Historical imports have no in-Cadence drafts to match. The user picks
     "+ Create new draft" (draft_id="__new__") on the rows and the commit
     endpoint creates a placeholder published draft per row.
"""
from __future__ import annotations

import base64
import io

import pytest


def _build_split_table_xlsx() -> bytes:
    """Synthesize an xlsx whose 'TOP POSTS' sheet mirrors LinkedIn's real
    layout: title row, blank, two subtables side-by-side separated by a
    blank column. URLs overlap between the subtables."""
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "TOP POSTS"

    ws.cell(row=1, column=1, value="Maximum of 50 posts available to include in this list.")

    # row 2 left blank intentionally

    # Headers on row 3 — two subtables, blank col D between them
    ws.cell(row=3, column=1, value="Post URL")
    ws.cell(row=3, column=2, value="Post publish date")
    ws.cell(row=3, column=3, value="Engagements")
    # col 4 left blank — the divider
    ws.cell(row=3, column=5, value="Post URL")
    ws.cell(row=3, column=6, value="Post publish date")
    ws.cell(row=3, column=7, value="Impressions")

    # Data — left table is sorted by engagement (rank: A, B, C),
    # right table is sorted by impressions (rank: B, A, C). The shared
    # URLs should let the merger combine engagements with impressions.
    data_left = [
        ("https://example.com/a", "4/19/2026", 98),
        ("https://example.com/b", "4/26/2026", 91),
        ("https://example.com/c", "4/18/2026", 45),
    ]
    data_right = [
        ("https://example.com/b", "4/26/2026", 4669),
        ("https://example.com/a", "4/19/2026", 2473),
        ("https://example.com/c", "4/18/2026", 2319),
    ]
    for i, (u, d, eng) in enumerate(data_left):
        ws.cell(row=4 + i, column=1, value=u)
        ws.cell(row=4 + i, column=2, value=d)
        ws.cell(row=4 + i, column=3, value=eng)
    for i, (u, d, impr) in enumerate(data_right):
        ws.cell(row=4 + i, column=5, value=u)
        ws.cell(row=4 + i, column=6, value=d)
        ws.cell(row=4 + i, column=7, value=impr)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_single_table_xlsx() -> bytes:
    """Backwards-compat case: an older / different LinkedIn export with a
    single straightforward table including all the breakdown columns."""
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Posts"
    ws.cell(row=1, column=1, value="Post URL")
    ws.cell(row=1, column=2, value="Created date")
    ws.cell(row=1, column=3, value="Impressions")
    ws.cell(row=1, column=4, value="Reactions")
    ws.cell(row=1, column=5, value="Comments")
    ws.cell(row=1, column=6, value="Reposts")
    rows = [
        ("https://example.com/a", "4/19/2026", 4669, 50, 30, 18),
        ("https://example.com/b", "4/26/2026", 2473, 60, 25, 6),
    ]
    for i, r in enumerate(rows):
        for j, v in enumerate(r):
            ws.cell(row=2 + i, column=j + 1, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Subtable detection + merge
# ---------------------------------------------------------------------------

def test_preview_merges_side_by_side_subtables(client, app_module):
    pytest.importorskip("openpyxl")
    xlsx = _build_split_table_xlsx()
    payload = {"xlsx": base64.b64encode(xlsx).decode("ascii")}
    r = client.post("/api/analytics/import-preview", json=payload)
    body = r.get_json()
    assert body["ok"] is True, body
    assert body["subtable_count"] == 2

    parsed = body["parsed"]
    # 3 unique URLs across the two subtables -> 3 merged rows, not 6.
    assert len(parsed) == 3
    by_url = {p["url"]: p for p in parsed}

    # Post A: engagements=98 (left table, maps into likes due to aggregate
    # fallback), impressions=2473 (right table). BEFORE this fix these
    # would have landed on different rows.
    assert by_url["https://example.com/a"]["likes"] == 98
    assert by_url["https://example.com/a"]["impressions"] == 2473

    # Post B: engagements=91, impressions=4669
    assert by_url["https://example.com/b"]["likes"] == 91
    assert by_url["https://example.com/b"]["impressions"] == 4669

    # Post C: engagements=45, impressions=2319
    assert by_url["https://example.com/c"]["likes"] == 45
    assert by_url["https://example.com/c"]["impressions"] == 2319


def test_preview_surfaces_aggregate_engagement_note(client, app_module):
    pytest.importorskip("openpyxl")
    xlsx = _build_split_table_xlsx()
    payload = {"xlsx": base64.b64encode(xlsx).decode("ascii")}
    r = client.post("/api/analytics/import-preview", json=payload)
    body = r.get_json()
    assert body["used_engagement_aggregate"] is True
    notes_joined = " ".join(body["notes"])
    assert "Engagements" in notes_joined or "engagement" in notes_joined.lower()
    assert "merged by URL" in notes_joined or "side-by-side" in notes_joined.lower()


def test_preview_single_table_unchanged(client, app_module):
    """Older single-table exports should still work — no subtable detection,
    no aggregate fallback, columns map straight through."""
    pytest.importorskip("openpyxl")
    xlsx = _build_single_table_xlsx()
    payload = {"xlsx": base64.b64encode(xlsx).decode("ascii")}
    r = client.post("/api/analytics/import-preview", json=payload)
    body = r.get_json()
    assert body["ok"] is True
    assert body["subtable_count"] == 1
    assert body["used_engagement_aggregate"] is False
    parsed = body["parsed"]
    assert len(parsed) == 2
    a = next(p for p in parsed if p["url"].endswith("/a"))
    assert a["impressions"] == 4669
    assert a["likes"] == 50
    assert a["comments"] == 30
    assert a["reposts"] == 18


# ---------------------------------------------------------------------------
# Auto-match should still work with subtable merge
# ---------------------------------------------------------------------------

def test_preview_still_auto_matches_against_existing_drafts(client, app_module):
    pytest.importorskip("openpyxl")
    with app_module.db_cursor() as conn:
        # Insert a draft whose body matches one of the URLs we'll see in
        # the import. The snippet matcher matches by body substring.
        conn.execute(
            "INSERT INTO drafts(title, body, status, posted_at) "
            "VALUES (?, ?, 'scheduled', '2026-04-19T00:00:00')",
            ("Existing draft", "https://example.com/a is a great post."),
        )
    xlsx = _build_split_table_xlsx()
    r = client.post(
        "/api/analytics/import-preview",
        json={"xlsx": base64.b64encode(xlsx).decode("ascii")},
    )
    body = r.get_json()
    matched = [p for p in body["parsed"] if p["matched_draft_id"]]
    # The merged record for /a has snippet="" because the URL was the only
    # field; date proximity should still pick it up (post date 4/19 matches
    # the seeded draft's posted_at).
    assert len(matched) >= 1
    assert any("a" in p["url"] for p in matched)


# ---------------------------------------------------------------------------
# Commit: __new__ creates published drafts for historical rows
# ---------------------------------------------------------------------------

def test_commit_creates_new_drafts_for_new_sentinel(client, app_module):
    payload = {"rows": [
        {
            "draft_id": "__new__",
            "date": "4/19/2026",
            "snippet": "",
            "url": "https://example.com/a",
            "impressions": 2473, "likes": 98,
            "comments": 0, "reposts": 0, "follows": 0,
        },
        {
            "draft_id": "__new__",
            "date": "4/26/2026",
            "snippet": "",
            "url": "https://example.com/b",
            "impressions": 4669, "likes": 91,
            "comments": 0, "reposts": 0, "follows": 0,
        },
        {
            # A skip-row mixed in
            "draft_id": None,
            "impressions": 9, "likes": 0, "comments": 0,
            "reposts": 0, "follows": 0,
        },
    ]}
    r = client.post("/api/analytics/import-commit", json=payload)
    body = r.get_json()
    assert body["ok"] is True
    assert body["created"] == 2
    assert body["skipped"] == 1
    assert len(body["new_drafts"]) == 2

    # Drafts should be status='published' with posted_at from the row date,
    # URL preserved in body so the user can still find the post on LinkedIn.
    drafts = client.get("/api/drafts").get_json()
    new_drafts = sorted(
        (d for d in drafts if d["status"] == "published"),
        key=lambda d: d["id"],
    )
    assert len(new_drafts) == 2
    a = new_drafts[0]
    assert "example.com/a" in a["body"]
    assert a["posted_at"].startswith("2026-04-19")

    # Analytics rows attached to the new drafts
    analytics = client.get("/api/analytics").get_json()
    assert len(analytics) == 2
    by_draft = {a["draft_id"]: a for a in analytics}
    for did in body["new_drafts"]:
        row = by_draft[did]
        assert row["impressions"] in (2473, 4669)
        assert row["likes"] in (98, 91)


def test_commit_title_falls_back_to_snippet_then_date(client, app_module):
    """Title heuristic order: explicit title > snippet > 'Imported · <date>'."""
    rows = [
        {  # Has snippet — use it
            "draft_id": "__new__",
            "date": "4/19/2026",
            "snippet": "This is the first 60 characters of the post body example",
            "url": "https://example.com/a",
            "impressions": 0, "likes": 0, "comments": 0,
            "reposts": 0, "follows": 0,
        },
        {  # No snippet, has date
            "draft_id": "__new__",
            "date": "4/26/2026",
            "snippet": "",
            "url": "https://example.com/b",
            "impressions": 0, "likes": 0, "comments": 0,
            "reposts": 0, "follows": 0,
        },
    ]
    r = client.post("/api/analytics/import-commit", json={"rows": rows})
    assert r.status_code == 200

    drafts = client.get("/api/drafts").get_json()
    new = [d for d in drafts if d["status"] == "published"]
    titles = {d["title"] for d in new}
    assert any(t.startswith("This is the first") for t in titles)
    assert any("Imported post · 4/26/2026" in t for t in titles)


def test_commit_handles_mix_of_new_and_existing(client, app_module):
    """One row matches an existing draft, one row creates a new one."""
    with app_module.db_cursor() as conn:
        cur = conn.execute(
            "INSERT INTO drafts(title, body, status) "
            "VALUES ('existing', 'pre-existing body', 'scheduled')"
        )
        existing_id = cur.lastrowid

    rows = [
        {
            "draft_id": existing_id,
            "impressions": 100, "likes": 10,
            "comments": 2, "reposts": 1, "follows": 0,
        },
        {
            "draft_id": "__new__",
            "date": "4/19/2026",
            "url": "https://example.com/historical",
            "impressions": 500, "likes": 25,
            "comments": 0, "reposts": 0, "follows": 0,
        },
    ]
    r = client.post("/api/analytics/import-commit", json={"rows": rows})
    body = r.get_json()
    assert body["ok"] is True
    assert body["created"] == 2
    assert body["skipped"] == 0
    assert len(body["new_drafts"]) == 1

    # The existing draft flipped to published
    drafts = client.get("/api/drafts").get_json()
    found = next(d for d in drafts if d["id"] == existing_id)
    assert found["status"] == "published"


def test_commit_rejects_invalid_draft_id_string(client, app_module):
    """Anything not numeric and not '__new__' is treated as skip, never
    crashes the loop."""
    r = client.post("/api/analytics/import-commit", json={"rows": [
        {"draft_id": "not_a_number",
         "impressions": 100, "likes": 0, "comments": 0,
         "reposts": 0, "follows": 0},
    ]})
    body = r.get_json()
    assert body["ok"] is True
    assert body["created"] == 0
    assert body["skipped"] == 1
